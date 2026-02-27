import requests
import json
import logging
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class StockIQAPI:
    def __init__(self):
        self.base_url = "https://earthrated.stockiqtech.net/api/"
        self.headers = {
            'Authorization': 'Basic WWluZzpLeDUmZn1IM0Y4',
            'Content-Type': 'application/json'
        }
        
        # Define filters
        self.site_filters = [
            "Northland Goreway",
            "Progressive UK",
            "Rhenus Netherlands",
            "Source Montebello",
            "Source NJ"
        ]
        
        self.supplier_filters = [
            "SUNNER GROUP CO., LTD.",
            "Tianjin Yiyi Hygiene Products Co., Ltd.",
            "DELON LAB. (1990) INC",
            "American Hygienics Corporation",
            "Thai Duong Rubber Joint Stock Company"
        ]
        
        # Define column mapping
        self.column_mapping = {
            'ItemCode': 'ItemCode',
            'SiteCode': 'Site',
            'TotalOnHandQuantity': 'On Hand Qty',
            'TotalOnOrderQuantity': 'On Order Qty',
            'SupplierName': 'Supplier',
            'ItemSiteCategory1Name': 'Category',
            'ItemSiteCategory4Name': 'NetSuite ABC',
            'InventoryPosition': 'InventoryPosition',
            'InventoryPositionAtLeadTime': 'InventoryPositionAtLeadTime',
            'WorstProjectedInventoryPosition': 'WorstProjectedInventoryPosition',
            'NextErpOrderNumber': 'Next Exp Order',
            'NextOrderRemainingReleaseQuantity': 'Next Exp Order Qty',
            'NextExpectedDockDate': 'NextExpectedDockDate',
            'NextExpectedShipDate': 'NextExpectedShipDate',
            'ProjectedBelowPanicPointDate': 'ProjectedBelowPanicPointDate',
            'ProjectedOutOfStockDate': 'ProjectedOutOfStockDate',
            'ActiveSafetyStock': 'Safety Stock',
            'ActiveMaxStock': 'Max Stock',
            'ActivePlanningLeadTime': 'Planning Lead Time'
        }
        
        # Define inventory status mapping
        self.inventory_status = {
            1: "Out of Stock",
            2: "Below Panic Point",
            4: "Below Safety Stock",
            8: "OK",
            24: "Overstocked",
            32: "Exceed",
            128: "Non Stock"
        }
    
    def fetch_order_schedule_header(self):
        """
        Fetch Order Schedule Header data from StockIQ API
        
        Returns:
            list: JSON data from API response
        """
        endpoint = "OrderScheduleHeader/"
        url = self.base_url + endpoint
        
        try:
            logger.info(f"Making API request to: {url}")
            response = requests.get(url, headers=self.headers, timeout=30)
            response.raise_for_status()
            
            data = response.json()
            logger.info(f"Successfully fetched {len(data)} records from API")
            return data
            
        except requests.exceptions.Timeout:
            logger.error("API request timed out")
            raise
        except requests.exceptions.ConnectionError:
            logger.error("Connection error when accessing API")
            raise
        except requests.exceptions.HTTPError as e:
            logger.error(f"HTTP error: {e}")
            logger.error(f"Response content: {response.text}")
            raise
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON response: {e}")
            raise
    
    def filter_data(self, data):
        """
        Filter data based on requirements
        
        Args:
            data (list): Raw data from API
            
        Returns:
            list: Filtered data
        """
        filtered_data = []
        
        for item in data:
            # Filter ItemCode starting with "10"
            if not str(item.get('ItemCode', '')).startswith('10'):
                continue
            
            # Filter SiteCode
            if item.get('SiteCode') not in self.site_filters:
                continue
            
            # Filter SupplierName
            if item.get('SupplierName') not in self.supplier_filters:
                continue
            
            filtered_data.append(item)
        
        logger.info(f"Filtered to {len(filtered_data)} records")
        return filtered_data
    
    def transform_data(self, data):
        """
        Transform data according to requirements
        
        Args:
            data (list): Filtered data
            
        Returns:
            pd.DataFrame: Transformed dataframe
        """
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Select only required columns
        required_columns = list(self.column_mapping.keys())
        
        # Keep only columns that exist in the data
        existing_columns = [col for col in required_columns if col in df.columns]
        df = df[existing_columns]
        
        # Rename columns
        df = df.rename(columns=self.column_mapping)
        
        # Transform inventory position columns
        inventory_cols = ['InventoryPosition', 'InventoryPositionAtLeadTime', 'WorstProjectedInventoryPosition']
        
        for col in inventory_cols:
            if col in df.columns:
                df[col] = df[col].map(self.inventory_status)
        
        logger.info(f"Transformed data with {len(df)} rows and {len(df.columns)} columns")
        return df
    
    def save_to_excel(self, df, filename=None):
        """
        Save dataframe to Excel with formatting
        
        Args:
            df (pd.DataFrame): Data to save
            filename (str): Optional filename
        """
        if df.empty:
            logger.warning("No data to save")
            return
        
        if filename is None:
            today = datetime.now().strftime("%Y-%m-%d")
            filename = f"StockIQ_OrderSchedule_Filtered_{today}.xlsx"
        
        filepath = Path(filename)
        
        try:
            logger.info(f"Saving {len(df)} records to {filepath}")
            
            # Save to Excel
            df.to_excel(filepath, index=False, sheet_name='Order Schedule')
            
            # Apply highlighting
            self.apply_highlighting(filepath, df)
            
            logger.info(f"Successfully saved data to {filepath}")
            logger.info(f"File size: {filepath.stat().st_size} bytes")
            
        except Exception as e:
            logger.error(f"Error writing to file {filepath}: {e}")
            raise
    
    def apply_highlighting(self, filepath, df):
        """
        Apply orange highlighting to rows where InventoryPositionAtLeadTime is "Below Panic Point"
        
        Args:
            filepath (Path): Path to Excel file
            df (pd.DataFrame): Original dataframe
        """
        try:
            # Load the workbook
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Define orange fill
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            
            # Get column indices
            itemcode_col = None
            site_col = None
            inventory_col = None
            
            for idx, col in enumerate(df.columns, start=1):
                if col == 'ItemCode':
                    itemcode_col = idx
                elif col == 'Site':
                    site_col = idx
                elif col == 'InventoryPositionAtLeadTime':
                    inventory_col = idx
            
            if not all([itemcode_col, site_col, inventory_col]):
                logger.warning("Could not find all required columns for highlighting")
                wb.save(filepath)
                return
            
            # Iterate through rows (starting from row 2, after header)
            for row_idx in range(2, len(df) + 2):
                cell_value = ws.cell(row=row_idx, column=inventory_col).value
                
                if cell_value == "Below Panic Point":
                    # Highlight ItemCode and Site columns
                    ws.cell(row=row_idx, column=itemcode_col).fill = orange_fill
                    ws.cell(row=row_idx, column=site_col).fill = orange_fill
            
            wb.save(filepath)
            logger.info("Applied highlighting to Excel file")
            
        except Exception as e:
            logger.error(f"Error applying highlighting: {e}")
            # Don't raise - file is still saved without highlighting
    
    def run(self):
        """
        Main execution method
        """
        try:
            # Fetch data from API
            data = self.fetch_order_schedule_header()
            
            # Filter data
            filtered_data = self.filter_data(data)
            
            if not filtered_data:
                logger.warning("No data after filtering")
                return
            
            # Transform data
            df = self.transform_data(filtered_data)
            
            # Save to Excel
            self.save_to_excel(df)
            
            logger.info("Process completed successfully")
            
        except Exception as e:
            logger.error(f"Process failed: {e}")
            raise

def main():
    """
    Entry point for the script
    """
    api_client = StockIQAPI()
    api_client.run()

if __name__ == "__main__":
    main()