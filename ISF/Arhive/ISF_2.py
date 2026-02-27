import gspread
import pandas as pd
import os
import io
import pickle
import base64
from datetime import timedelta

# --- Third-party libraries ---
from openpyxl import load_workbook
from gspread_dataframe import get_as_dataframe
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage # <-- Import MIMEImage for the signature logo

# ==============================================================================
# SCRIPT CONFIGURATION - - MODIFY THIS SECTION
# ==============================================================================

INTERNAL_IDS_TO_PROCESS = ['2045']
CLIENT_SECRETS_FILE = 'client_secret_445547975943-l4bvv3advto71sovc6rn5ndai1gmhib9.apps.googleusercontent.com.json'
PO_CHECK_SPREADSHEET_ID = '1DDb_Z08ET94wdmD6a93UPVgWQ73y0F2tQIlgHqMiFCc'
INBOUND_SHIPMENT_TRACKER_TAB = 'Inbound Shipment Tracker'
SENDER_EMAIL = "michelle@earthrated.com"
TO_RECIPIENTS = ["customs.phl@savinodelbene.com, kadyn.pease@savinodelbene.com"]
CC_RECIPIENTS = ["michelle@earthrated.com"]
ISF_TEMPLATE_FILENAME = 'ISF_Python.xlsx'

# --- NEW: Signature Configuration ---
SIGNATURE_IMAGE_FILENAME = 'signature_logo.png' # The logo file in the same directory

# ==============================================================================
# AUTHENTICATION FUNCTIONS (Do not modify)
# ==============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SHEETS_SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']

def get_google_creds(token_name, scopes):
    token_path = os.path.join(SCRIPT_DIR, token_name)
    creds = None
    if os.path.exists(token_path):
        with open(token_path, 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, scopes)
            creds = flow.run_local_server(port=0)
        with open(token_path, 'wb') as token:
            pickle.dump(creds, token)
    return creds

# ==============================================================================
# CORE FUNCTIONS (Do not modify unless specified)
# ==============================================================================

def fill_isf_form(filename, data):
    # This function remains unchanged
    wb = load_workbook(filename)
    ws = wb.active
    po_numbers = data.get('purchase_order_number', '').split(';')
    for index, po in enumerate(po_numbers):
        if po: ws[f'{chr(71+index)}5'] = po.strip()
    ws['G6'] = data.get('sailing_date', '')
    ws['G8'] = data.get('mbl_scace_code', '')
    ws['G9'] = data.get('container_number', '')
    ws['G10'] = data.get('vessel_number', '')
    ws['G13'] = data.get('ams_house_bills', '')
    ws['D14'] = data.get('hbl_scac_code', '')
    manufacturer = data.get('manufacturer', '')
    if manufacturer == "American Hygienics Corporation":
        for cell in ['H18', 'H30', 'H42', 'H48']: ws[cell] = "AHC LTD."
        for cell in ['H19', 'H31', 'H43', 'H49']: ws[cell] = "189,HAIZHEN RD,DUSHANGANG TOWN, PINGHU"
        for cell in ['H21', 'H33', 'H45', 'H51']: ws[cell] = "JIAXING"
        for cell in ['H22', 'H34', 'H46', 'H52']: ws[cell] = "ZHEJIANG"
        for cell in ['H23', 'H35', 'H47', 'H53']: ws[cell] = "CHINA"
        ws['G25'] = "3401.19"
    elif manufacturer == "SUNNER GROUP CO., LTD.":
        for addr in ['H18', 'H42']: ws[addr] = "Yancheng Chong Jia Technology Co., Ltd"
        for addr in ['H19', 'H43']: ws[addr] = "West side of youdian road, Qinfeng mechanical and electrical industrial park,"
        for city in ['H20', 'H44']: ws[city] = "Jianhu town"
        for province in ['H21', 'H45']: ws[province] = "Yan Cheng"
        for province in ['H22', 'H46']: ws[province] = "Jiang Su"
        for country in ['H23', 'H29', 'H35', 'H47', 'H53']: ws[country] = "CHINA"
        ws['G25'] = "3923.21"
        for addr in ['H30', 'H48']: ws[addr] = "Sunner World Trading CO., Ltd."
        for addr in ['H31', 'H49']: ws[addr] = "Suite 706-707 Huitu Mansion 898 Tong Pu Road"
        for city in ['H33', 'H51', 'H34', 'H52']: ws[city] = "Shanghai"
    elif manufacturer == "Thai Duong Rubber Joint Stock Company":
        for cell in ['H18', 'H30', 'H42', 'H48']: ws[cell] = "THAI DUONG RUBBER JOINT STOCK COMPANY"
        for cell in ['H19', 'H31', 'H43', 'H49']: ws[cell] = "LOT 4, ROAD 7, TAN TAO INDUSTRIAL PARK, BINH TAN DISTRICT"
        for cell in ['H21', 'H33', 'H45', 'H51']: ws[cell] = "HO CHI MINH CITY"
        for cell in ['H22', 'H34', 'H46', 'H52']: ws[cell] = "HO CHI MINH CITY"
        for cell in ['H23', 'H35', 'H47', 'H53']: ws[cell] = "VIETNAM"
        ws['G25'] = "4016.99"
    elif manufacturer == "Tianjin Yiyi Hygiene Products Co., Ltd.":
        for cell in ['H18', 'H30', 'H42', 'H48']: ws[cell] = "TIANJIN YIYI HYGIENE PRODUCTS CO., LTD."
        for cell in ['H19', 'H31', 'H43', 'H49']: ws[cell] = "ZHANGJIAWO INDUSTRIAL PARK,XIQING ECONOMIC DEVELOPMENT AREA"
        for cell in ['H21', 'H33', 'H45', 'H51']: ws[cell] = "TIAN JIN"
        for cell in ['H22', 'H34', 'H46', 'H52']: ws[cell] = "TIAN JIN"
        for cell in ['H23', 'H35', 'H47', 'H53']: ws[cell] = "CHINA"
        ws['G25'] = "4818.90"
    destination_warehouse = data.get('destination_warehouse', '')
    if destination_warehouse == "Source NJ":
        ws['H54'], ws['H55'], ws['H57'], ws['H58'] = "SOURCE NEW JERSEY", "2 MIDDLESEX AVENUE, SUITE A", "MONROE", "NJ"
    elif destination_warehouse == "Source Montebello":
        ws['H54'], ws['H55'], ws['H56'], ws['H57'] = "SOURCE LOGISTICS LOS ANGELES, LLC", "812 UNION ST", "MONTEBELLO", "CA"
    po_string = '+'.join([po.strip() for po in po_numbers if po])
    new_filename = f"{po_string} ISF.xlsx"
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return new_filename, output_stream.getvalue()

# --- UPDATED: This function now sends HTML emails with an embedded image ---
def send_email_with_attachment(to, cc, subject, html_body, sender, attachment_name, attachment_data, image_path):
    """Sends an HTML email with an attachment and embedded image using the Gmail API."""
    creds = get_google_creds('token_gmail.pickle', GMAIL_SCOPES)
    service = build('gmail', 'v1', credentials=creds)
    
    # Create a 'related' multipart message. This is key for embedding images.
    message = MIMEMultipart('related')
    message['to'] = ", ".join(to)
    if cc:
        message['cc'] = ", ".join(cc)
    message['from'] = sender
    message['subject'] = subject

    # Create the HTML part of the message
    msg_alternative = MIMEMultipart('alternative')
    message.attach(msg_alternative)
    msg_text = MIMEText(html_body, 'html')
    msg_alternative.attach(msg_text)

    # Embed the signature image
    try:
        with open(image_path, 'rb') as f:
            msg_image = MIMEImage(f.read())
            # Add a 'Content-ID' header. The 'cid' in the HTML must match this.
            msg_image.add_header('Content-ID', '<signature_logo>')
            message.attach(msg_image)
    except FileNotFoundError:
        print(f"Warning: Signature image '{image_path}' not found. Sending email without logo.")

    # Attach the ISF file
    part = MIMEApplication(attachment_data, Name=attachment_name)
    part['Content-Disposition'] = f'attachment; filename="{attachment_name}"'
    message.attach(part)
    
    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
    try:
        sent_message = service.users().messages().send(userId='me', body={'raw': raw_message}).execute()
        all_recipients = to + (cc if cc else [])
        print(f"Email with attachment sent successfully to {all_recipients}. Message ID: {sent_message['id']}")
    except Exception as error:
        print(f"An error occurred while sending email: {error}")

# ==============================================================================
# MAIN EXECUTION SCRIPT
# ==============================================================================
if __name__ == "__main__":
    print("Starting ISF automation script...")
    
    try:
        sheets_creds = get_google_creds('token_sheets.pickle', SHEETS_SCOPES)
        client = gspread.authorize(sheets_creds)
        sheet = client.open_by_key(PO_CHECK_SPREADSHEET_ID)
        worksheet = sheet.worksheet(INBOUND_SHIPMENT_TRACKER_TAB)
        df = get_as_dataframe(worksheet, evaluate_formulas=True)
        df['Internal Id'] = df['Internal Id'].astype(str).str.strip().replace(r'\.0$', '', regex=True)
        print("Successfully loaded data from 'Inbound Shipment Tracker'.")
    except Exception as e:
        print(f"Error: Could not load Google Sheet. Details: {e}")
        exit()

    for internal_id in INTERNAL_IDS_TO_PROCESS:
        print(f"\n--- Processing Internal ID: {internal_id} ---")
        
        shipment_df = df[df['Internal Id'] == internal_id].copy()
        
        if shipment_df.empty:
            print(f"Warning: Internal ID {internal_id} not found. Skipping.")
            continue

        first_row = shipment_df.iloc[0]
        
        po_numbers = ";".join(shipment_df['PO#'].astype(str).unique())
        mbl_number = str(first_row.get('MBL#', ''))
        manufacturer = str(first_row.get('Supplier Name', ''))
        ams_house_bills, hbl_scac_code = "", ""
        if not mbl_number.upper().startswith('COSU'):
            ams_house_bills = str(first_row.get('Booking#', ''))
            if ams_house_bills and len(ams_house_bills) >= 4:
                hbl_scac_code = ams_house_bills[:4]

        isf_data = {
            'purchase_order_number': po_numbers,
            'sailing_date': pd.to_datetime(first_row['ETD']).strftime('%Y-%m-%d') if pd.notna(first_row['ETD']) else '',
            'mbl_scace_code': mbl_number,
            'container_number': str(first_row.get('Container#', '')),
            'vessel_number': '',
            'ams_house_bills': ams_house_bills,
            'hbl_scac_code': hbl_scac_code,
            'manufacturer': manufacturer,
            'destination_warehouse': str(first_row.get('Location', ''))
        }
        print("Successfully extracted data for ISF form.")

        try:
            generated_filename, attachment_data = fill_isf_form(ISF_TEMPLATE_FILENAME, isf_data)
            print(f"Successfully generated ISF file in memory: {generated_filename}")
        except FileNotFoundError:
            print(f"Error: The template file '{ISF_TEMPLATE_FILENAME}' was not found.")
            continue
        except Exception as e:
            print(f"Error during ISF form generation: {e}")
            continue

        po_subject_part = "+".join(po_numbers.split(';'))
        shipping_port_map = {"American Hygienics Corporation": "SHA", "SUNNER GROUP CO., LTD.": "SHA", "Tianjin Yiyi Hygiene Products Co., Ltd.": "XNG", "Thai Duong Rubber Joint Stock Company": "VNM"}
        shipping_port = shipping_port_map.get(manufacturer, "N/A")
        etd_date_str = pd.to_datetime(first_row['ETD']).strftime('%-m/%-d') if pd.notna(first_row['ETD']) else 'N/A'
        eta_date_str = 'N/A'
        if pd.notna(first_row['Est Delivery Date']):
            eta_date = pd.to_datetime(first_row['Est Delivery Date']) - timedelta(days=5)
            eta_date_str = eta_date.strftime('%-m/%-d')
        location = str(first_row.get('Location', ''))
        email_subject = f"EARTH RATED ISF filing - {internal_id} - {po_subject_part} ETD {shipping_port} {etd_date_str} ETA {location} {eta_date_str}"
        
        # --- NEW: Define the HTML content for the email body and signature ---
        html_content = """
        <p>Hi Team,</p>
        <p>Good Day!</p>
        <p>Pls help to file ISF for subject PO.</p>
        <p>Let me know if there is any issues.</p>
        <p>Thanks a lot. Have a great day!</p>
        <br>
        <table cellpadding="0" cellspacing="0" border="0" style="font-family: Arial, sans-serif; font-size: 10pt; color: #333333;">
            <tr>
                <td valign="top" style="padding-right: 15px;">
                    <img src="cid:signature_logo" alt="Earth Rated Logo" width="120">
                </td>
                <td valign="top" style="border-left: 1px solid #dcdcdc; padding-left: 15px; font-size: 10pt;">
                    <b style="color: #00573c; font-size: 11pt;">Michelle Zhao</b><br>
                    Supply Planner<br>
                    Planificateur de l'approvisionnement<br><br>
                    1.888.354.2818<br>
                    earthrated.com
                </td>
            </tr>
        </table>
        """

        print(f"Preparing to send email with subject: {email_subject}")
        send_email_with_attachment(
            to=TO_RECIPIENTS,
            cc=CC_RECIPIENTS,
            subject=email_subject,
            html_body=html_content, # Pass the HTML content
            sender=SENDER_EMAIL,
            attachment_name=generated_filename,
            attachment_data=attachment_data,
            image_path=os.path.join(SCRIPT_DIR, SIGNATURE_IMAGE_FILENAME) # Pass the image path
        )

    print("\n--- Script finished. ---")