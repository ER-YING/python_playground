import requests
import pandas as pd
from datetime import datetime
import json
import os
import re
from typing import List
import warnings
import openpyxl
from openpyxl.styles import PatternFill, Border, Side

# Suppress the specific FutureWarning for concat
warnings.filterwarnings('ignore', category=FutureWarning, message='.*DataFrame concatenation.*')

# --------- Step 1: API and Data Processing Functions (Keep same as before) ---------

def get_stockiq_custom_report():
    # Configuration
    BASE_URL = 'https://earthrated.stockiqtech.net/api/'
    CUSTOM_REPORT_ID = 4
    AUTHORIZATION = 'Basic WWluZzpLeDUmZn1IM0Y4'
    
    # API endpoint
    endpoint = f"{BASE_URL}CustomReportProducer"
    
    # Headers
    headers = {
        'Authorization': AUTHORIZATION,
        'Content-Type': 'application/json',
        'Accept': 'application/json'
    }
    
    # Parameters
    params = {
        'customReportId': CUSTOM_REPORT_ID
    }
    
    try:
        print("Making API call to StockIQ...")
        response = requests.get(endpoint, headers=headers, params=params)
        
        if response.status_code == 200:
            data = response.json()
            print(f"Successfully retrieved {len(data)} records")
            
            df = pd.DataFrame(data)
            
            if not df.empty:
                print("\nAvailable fields in the report:")
                for i, field in enumerate(df.columns, 1):
                    print(f"{i:2d}. {field}")
                
                return df
            else:
                print("No data returned from the API")
                return None
                
        else:
            print(f"Error: API call failed with status code {response.status_code}")
            print(f"Response: {response.text}")
            return None
            
    except Exception as e:
        print(f"Error: {e}")
        return None

def get_dependent_forecast_data():
    """Get Dependent Forecast report from StockIQ"""
    BASE_URL = 'https://earthrated.stockiqtech.net/api/'
    CUSTOM_REPORT_NAME = 'Dependent Forecast'
    AUTHORIZATION = 'Basic WWluZzpLeDUmZn1IM0Y4'
    
    endpoint = f"{BASE_URL}CustomReportProducer"
    
    headers = {
        'Authorization': AUTHORIZATION,
        'Content-Type': 'application/json',
        'Accept': 'application/json'
    }
    
    params = {
        'customReportName': CUSTOM_REPORT_NAME
    }
    
    try:
        print("Fetching Dependent Forecast report...")
        response = requests.get(endpoint, headers=headers, params=params)
        
        if response.status_code == 200:
            data = response.json()
            df = pd.DataFrame(data)
            print(f"Successfully retrieved Dependent Forecast: {len(df)} records")
            return df
        else:
            print(f"Error fetching Dependent Forecast: {response.status_code}")
            return pd.DataFrame()  # Return empty DataFrame if failed
            
    except Exception as e:
        print(f"Error fetching Dependent Forecast: {e}")
        return pd.DataFrame()  # Return empty DataFrame if failed

def apply_filters(df):
    """Apply filters to the dataframe - STRICT FILTERING"""
    if df is None or df.empty:
        return df
    
    print(f"\n🔍 APPLYING STRICT FILTERS...")
    print(f"Original record count: {len(df)}")
    
    # Define filters - EXACTLY as specified
    target_sites = {
        "Source NJ", "Northland Goreway", "Source Montebello",
        "Progressive UK", "Rhenus Netherlands", "Golden State FC LLC (AGL)", "Yanada"
    }
    
    target_categories = {
        "Cleanup : Bags", "Cleanup : Dispensers", "Cleanup : Pee Pads", "Grooming : Wipes",
        "Grooming : Specialty Wipes", "Grooming : Waterless",
        "Grooming : Shampoo", "Toys : Interactive Toys", "Toys : Standalone"
    }
    
    # Item filter - items starting with "10"
    item_prefix = "10"
    
    # Show what's actually in the data BEFORE filtering
    print("\n" + "="*60)
    print("📋 CURRENT DATA ANALYSIS:")
    print("="*60)
    
    if 'SiteCode' in df.columns:
        unique_sites = sorted(df['SiteCode'].unique())
        print(f"🏢 ALL SITES IN DATA ({len(unique_sites)}):")
        for site in unique_sites:
            count = len(df[df['SiteCode'] == site])
            in_target = "✅" if site in target_sites else "❌"
            print(f"   {in_target} '{site}' ({count} records)")
    
    if 'ProductCategory' in df.columns:
        unique_categories = sorted(df['ProductCategory'].unique())
        print(f"\n📂 ALL CATEGORIES IN DATA ({len(unique_categories)}):")
        for category in unique_categories:
            count = len(df[df['ProductCategory'] == category])
            in_target = "✅" if category in target_categories else "❌"
            print(f"   {in_target} '{category}' ({count} records)")
    
    if 'ItemCode' in df.columns:
        # Show sample of item codes and count with prefix
        items_with_prefix = df[df['ItemCode'].astype(str).str.startswith(item_prefix)]
        items_without_prefix = df[~df['ItemCode'].astype(str).str.startswith(item_prefix)]
        print(f"\n🏷️  ITEM CODE ANALYSIS:")
        print(f"   ✅ Items starting with '{item_prefix}': {len(items_with_prefix)} records")
        print(f"   ❌ Items NOT starting with '{item_prefix}': {len(items_without_prefix)} records")
        
        # Show examples
        if len(items_with_prefix) > 0:
            print(f"   Examples with '{item_prefix}': {items_with_prefix['ItemCode'].head(5).tolist()}")
        if len(items_without_prefix) > 0:
            print(f"   Examples without '{item_prefix}': {items_without_prefix['ItemCode'].head(5).tolist()}")
    
    print("\n" + "="*60)
    print("🔄 APPLYING FILTERS:")
    print("="*60)
    
    # Start with original data
    filtered_df = df.copy()
    
    # Filter 1: Sites
    if 'SiteCode' in df.columns:
        before_count = len(filtered_df)
        filtered_df = filtered_df[filtered_df['SiteCode'].isin(target_sites)]
        after_count = len(filtered_df)
        print(f"🏢 Site filter: {before_count} → {after_count} records")
        
        if after_count == 0:
            print(f"   ⚠️  NO SITES MATCHED! Check if site names match exactly.")
        else:
            matched_sites = sorted(filtered_df['SiteCode'].unique())
            print(f"   ✅ Matched sites: {matched_sites}")
    
    # Filter 2: Categories (only if we still have data)
    if 'ProductCategory' in df.columns and len(filtered_df) > 0:
        before_count = len(filtered_df)
        filtered_df = filtered_df[filtered_df['ProductCategory'].isin(target_categories)]
        after_count = len(filtered_df)
        print(f"📂 Category filter: {before_count} → {after_count} records")
        
        if after_count == 0:
            print(f"   ⚠️  NO CATEGORIES MATCHED! Check if category names match exactly.")
        else:
            matched_categories = sorted(filtered_df['ProductCategory'].unique())
            print(f"   ✅ Matched categories: {matched_categories}")
    
    # Filter 3: Item prefix (only if we still have data)
    if 'ItemCode' in df.columns and len(filtered_df) > 0:
        before_count = len(filtered_df)
        filtered_df = filtered_df[filtered_df['ItemCode'].astype(str).str.startswith(item_prefix)]
        after_count = len(filtered_df)
        print(f"🏷️  Item prefix filter: {before_count} → {after_count} records")
        
        if after_count == 0:
            print(f"   ⚠️  NO ITEMS MATCHED prefix '{item_prefix}'!")
        else:
            print(f"   ✅ Items with prefix '{item_prefix}' retained")
    
    print("\n" + "="*60)
    print(f"📊 FINAL RESULT: {len(filtered_df)} records after all filters")
    print("="*60)
    
    if len(filtered_df) == 0:
        print("🚨 NO DATA REMAINS AFTER FILTERING!")
        print("💡 This means your filter criteria don't match the actual data values.")
        print("💡 Check the analysis above to see what values actually exist.")
        
        # Offer to continue with original data
        continue_choice = input("\n❓ Do you want to proceed with unfiltered data? (y/N): ").strip().lower()
        if continue_choice == 'y':
            print("📝 Proceeding with original data...")
            return df
    
    return filtered_df

def save_mrp_csv(df):
    """Save the MRP report as CSV with today's date"""
    if df is None or df.empty:
        print("❌ No data to save")
        return None
    
    print(f"\n📊 Saving MRP report with ALL FIELDS: {list(df.columns)}")
    
    # Generate filename
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"MRP_{today}.csv"
    
    # Save to CSV
    try:
        df.to_csv(filename, index=False)
        print(f"\n✅ MRP Report saved successfully!")
        print(f"📁 Filename: {filename}")
        print(f"📍 Location: {os.path.abspath(filename)}")
        print(f"📊 Records: {len(df)}")
        print(f"🔧 Fields: {len(df.columns)} - {list(df.columns)}")
        
        # Show a preview of the data
        if len(df) > 0:
            print(f"\n👀 Preview (first 3 rows):")
            print(df.head(3).to_string(index=False))
        
        return filename
        
    except Exception as e:
        print(f"❌ Error saving file: {e}")
        return None

# --------- Step 2: Target Days Processing Functions ---------

def normalize_key(x):
    """Normalize strings for matching keys."""
    if pd.isna(x):
        return ""
    s = str(x).replace("\u00A0", " ").strip()       # remove non-breaking spaces
    s = re.sub(r"\s+", " ", s)                      # collapse multiple spaces
    return s.upper()

def find_col(cols: List[str], preferred_exact: List[str], contains_any: List[str]):
    """Find column name in a DataFrame that matches preferred names."""
    lower = {c.lower(): c for c in cols}
    for name in preferred_exact:
        if name.lower() in lower:
            return lower[name.lower()]
    for c in cols:
        cl = c.lower()
        if any(pat.lower() in cl for pat in contains_any):
            return c
    raise KeyError(f"Could not find matching column in {cols}")

def get_excel_column_name(col_index):
    """Convert 0-based column index to Excel column name (0->A, 1->B, etc.)"""
    result = ""
    while col_index >= 0:
        result = chr(col_index % 26 + ord('A')) + result
        col_index = col_index // 26 - 1
        if col_index < 0:
            break
    return result

def extract_month_year_from_column(column_name):
    """Extract MM/YYYY from column names like 'Aug/2025', 'Sep/2025', etc."""
    try:
        # Handle formats like "Aug/2025", "Sep/2025"
        if '/' in column_name:
            month_str, year_str = column_name.split('/')
            month_str = month_str.strip()
            year_str = year_str.strip()
            
            # Convert month name to number
            month_map = {
                'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
                'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
                'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12',
                'january': '01', 'february': '02', 'march': '03', 'april': '04',
                'june': '06', 'july': '07', 'august': '08', 'september': '09',
                'october': '10', 'november': '11', 'december': '12'
            }
            
            month_num = month_map.get(month_str.lower(), month_str)
            return f"{month_num}/{year_str}"
    except:
        pass
    
    return None

def create_dependent_forecast_lookup(dependent_forecast_df, month_columns):
    """Create lookup dictionary for dependent forecast data with better month matching"""
    lookup = {}
    
    if dependent_forecast_df.empty:
        return lookup
    
    # Try different possible column names
    item_col_names = ['ItemCode', 'Item Code', 'item_code', 'Item_Code']
    site_col_names = ['SiteCode', 'Site Code', 'site_code', 'Site_Code']
    date_col_names = ['MonthDate', 'Month Date', 'month_date', 'Month_Date', 'Date']
    forecast_col_names = ['DependentForecast', 'Dependent Forecast', 'dependent_forecast', 'Dependent_Forecast', 'Forecast']
    
    # Find the actual column names
    item_col = None
    site_col = None
    date_col = None
    forecast_col = None
    
    for col in dependent_forecast_df.columns:
        if col in item_col_names:
            item_col = col
        elif col in site_col_names:
            site_col = col
        elif col in date_col_names:
            date_col = col
        elif col in forecast_col_names:
            forecast_col = col
    
    if not all([item_col, site_col, forecast_col]):
        print(f"⚠️ Could not find required columns in Dependent Forecast data")
        print(f"Available columns: {list(dependent_forecast_df.columns)}")
        return lookup
    
    print(f"📋 Using Dependent Forecast columns: Item={item_col}, Site={site_col}, Date={date_col}, Forecast={forecast_col}")
    
    # Create lookup with better month matching
    for _, row in dependent_forecast_df.iterrows():
        item_key = normalize_key(row[item_col])
        site_key = normalize_key(row[site_col])
        forecast_value = row[forecast_col] if pd.notna(row[forecast_col]) else 0
        
        if date_col and pd.notna(row[date_col]):
            # Convert MM/DD/YYYY to MM/YYYY for matching
            try:
                date_str = str(row[date_col])
                # Handle different date formats
                if '/' in date_str:
                    parts = date_str.split('/')
                    if len(parts) >= 2:
                        month = parts[0].zfill(2)  # Ensure 2 digits
                        year = parts[-1]  # Take last part as year
                        month_year = f"{month}/{year}"
                        
                        key = (item_key, site_key, month_year)
                        lookup[key] = forecast_value
                        
                        # Debug print for verification
                        if item_key.startswith('10CA10BG0002') and site_key.startswith('SOURCE NJ'):
                            print(f"📝 Added lookup: {key} = {forecast_value}")
            except Exception as e:
                print(f"⚠️ Error processing date {row[date_col]}: {e}")
                continue
    
    print(f"📊 Created {len(lookup)} dependent forecast lookup entries")
    return lookup

def create_enhanced_rows(group, month_columns, days_value, starting_row_num, dependent_forecast_df):
    """Create enhanced rows with better Dependent Forecast matching"""
    
    # Get the basic info from the first row
    head = group.iloc[0]
    
    # Create a dictionary to store step data
    step_data = {}
    
    # Process existing rows
    for _, row in group.iterrows():
        step_data[row['Step']] = row
    
    # Define the correct order with new steps and their row positions
    ordered_steps = [
        ("1 Beginning On Hand", 0),      # Row offset 0
        ("2 Demand", 1),                 # Row offset 1  
        ("Simulated demand", 2),         # Row offset 2 (NEW)
        ("3 Forecast", 3),               # Row offset 3
        ("Dependent Forecast", 4),       # Row offset 4 (NEW)
        ("4 PO On Order", 5),            # Row offset 5
        ("5 Shipment", 6),               # Row offset 6
        ("6 Suggested PO Receipt", 7),   # Row offset 7
        ("Simulated Order", 8),          # Row offset 8 (NEW)
        ("7 Ending Inventory", 9),       # Row offset 9
        ("8 DOC - End of Month", 10),    # Row offset 10
        ("9 TDOC", 11)                   # Row offset 11
    ]
    
    # Create lookup for dependent forecast data
    dependent_forecast_lookup = create_dependent_forecast_lookup(dependent_forecast_df, month_columns)
    
    # Create ordered rows list
    ordered_rows = []
    
    # Find the column index where month data starts (after "Step")
    step_col_index = None
    for i, col in enumerate(group.columns):
        if col == 'Step':
            step_col_index = i
            break
    
    first_month_col_index = step_col_index + 1 if step_col_index is not None else 4
    
    for step, row_offset in ordered_steps:
        current_row_num = starting_row_num + row_offset
        
        if step in step_data:
            # Use existing data
            row_dict = step_data[step].to_dict()
            ordered_rows.append(row_dict)
        elif step == "Simulated demand":
            # Create new Simulated demand row
            sim_demand_row = step_data["2 Demand"].to_dict().copy()
            sim_demand_row['Step'] = "Simulated demand"
            # Set all month values to 0
            for mc in month_columns:
                sim_demand_row[mc] = 0
            ordered_rows.append(sim_demand_row)
        elif step == "Dependent Forecast":
            # Create new Dependent Forecast row
            dependent_row = step_data["3 Forecast"].to_dict().copy()
            dependent_row['Step'] = "Dependent Forecast"
            
            # Look up dependent forecast values for each month
            item_key = normalize_key(head['ItemCode'])
            site_key = normalize_key(head['SiteCode'])
            
            print(f"🔍 Looking up Dependent Forecast for: {item_key} - {site_key}")
            
            for mc in month_columns:
                forecast_value = 0  # Default value
                
                # Extract MM/YYYY from the column name (e.g., "Aug/2025" -> "08/2025")
                month_year = extract_month_year_from_column(mc)
                
                if month_year:
                    # Look for exact match
                    key = (item_key, site_key, month_year)
                    if key in dependent_forecast_lookup:
                        forecast_value = dependent_forecast_lookup[key]
                        print(f"✅ Found match for {mc} ({month_year}): {forecast_value}")
                    else:
                        print(f"❌ No match found for {mc} ({month_year})")
                
                dependent_row[mc] = forecast_value
            ordered_rows.append(dependent_row)
        elif step == "Simulated Order":
            # Create new Simulated Order row  
            sim_order_row = step_data["6 Suggested PO Receipt"].to_dict().copy()
            sim_order_row['Step'] = "Simulated Order"
            # Set all month values to 0
            for mc in month_columns:
                sim_order_row[mc] = 0
            ordered_rows.append(sim_order_row)
    
    # Create DataFrame from ordered rows
    enhanced_df = pd.DataFrame(ordered_rows)
    
    # Convert month columns to object type to allow formulas
    for mc in month_columns:
        if mc in enhanced_df.columns:
            enhanced_df[mc] = enhanced_df[mc].astype('object')
    
    # Apply formulas for specific steps
    for i, row in enhanced_df.iterrows():
        step = row['Step']
        current_excel_row = starting_row_num + i + 1  # +1 because Excel rows are 1-based
        
        # Update Ending Inventory with formula (updated to include Dependent Forecast)
        if step == "7 Ending Inventory":
            for col_idx, mc in enumerate(month_columns):
                excel_col = get_excel_column_name(first_month_col_index + col_idx)
                
                # Calculate absolute row references within this group (updated positions)
                beginning_row = starting_row_num + 0      # "1 Beginning On Hand"
                demand_row = starting_row_num + 1         # "2 Demand"
                sim_demand_row = starting_row_num + 2     # "Simulated demand"
                forecast_row = starting_row_num + 3       # "3 Forecast"
                dependent_forecast_row = starting_row_num + 4  # "Dependent Forecast"
                po_order_row = starting_row_num + 5       # "4 PO On Order"
                shipment_row = starting_row_num + 6       # "5 Shipment"
                suggested_po_row = starting_row_num + 7   # "6 Suggested PO Receipt"
                sim_order_row = starting_row_num + 8      # "Simulated Order"
                
                # Formula: Beginning + PO + Shipment + Suggested + SimOrder - Demand - SimDemand - Forecast - DependentForecast
                formula = f"={excel_col}{beginning_row}+{excel_col}{po_order_row}+{excel_col}{shipment_row}+{excel_col}{suggested_po_row}+{excel_col}{sim_order_row}-{excel_col}{demand_row}-{excel_col}{sim_demand_row}-{excel_col}{forecast_row}-{excel_col}{dependent_forecast_row}"
                enhanced_df.at[i, mc] = formula
        
        # Update DOC - End of Month with formula (updated to include Dependent Forecast in calculation)
        elif step == "8 DOC - End of Month":
            for col_idx, mc in enumerate(month_columns):
                excel_col = get_excel_column_name(first_month_col_index + col_idx)
                ending_inventory_row = starting_row_num + 9  # "7 Ending Inventory" row (updated position)
                
                # Calculate next 3 months columns
                next_3_months_cols = []
                for j in range(1, 4):  # Next 3 months
                    if col_idx + j < len(month_columns):
                        next_col = get_excel_column_name(first_month_col_index + col_idx + j)
                        next_3_months_cols.append(next_col)
                    else:
                        next_col = get_excel_column_name(first_month_col_index + len(month_columns) - 1)
                        next_3_months_cols.append(next_col)
                
                # Build the sum range for next 3 months' demand (updated row positions)
                demand_row = starting_row_num + 1         # "2 Demand"
                sim_demand_row = starting_row_num + 2     # "Simulated demand"
                forecast_row = starting_row_num + 3       # "3 Forecast"
                dependent_forecast_row = starting_row_num + 4  # "Dependent Forecast"
                
                # Create sum formula for next 3 months (including Dependent Forecast)
                sum_parts = []
                for next_col in next_3_months_cols:
                    sum_parts.append(f"{next_col}{demand_row}+{next_col}{sim_demand_row}+{next_col}{forecast_row}+{next_col}{dependent_forecast_row}")
                
                sum_formula = "+".join(sum_parts)
                
                # Final DOC formula
                formula = f"=IFERROR({excel_col}{ending_inventory_row}/(({sum_formula})/90),\"\")"
                enhanced_df.at[i, mc] = formula

    # Update Beginning On Hand for months 2+ after all formulas are set
    for i, row in enhanced_df.iterrows():
        if row['Step'] == "1 Beginning On Hand":
            for col_idx, mc in enumerate(month_columns):
                if col_idx > 0:  # Skip first month (keep original value)
                    prev_excel_col = get_excel_column_name(first_month_col_index + col_idx - 1)
                    ending_inventory_row = starting_row_num + 9  # "7 Ending Inventory" row (updated position)
                    
                    # Set Beginning On Hand = previous month's Ending Inventory
                    beginning_formula = f"={prev_excel_col}{ending_inventory_row}"
                    enhanced_df.at[i, mc] = beginning_formula
            break
    
    # Add TDOC row at the end
    tdoc_row = {}
    for col in enhanced_df.columns:
        if col in ['ProductCategory', 'ItemCode', 'SiteCode']:
            tdoc_row[col] = head[col]
        elif col == 'Step':
            tdoc_row[col] = "9 TDOC"
        elif col in month_columns:
            tdoc_row[col] = days_value
        else:
            tdoc_row[col] = pd.NA
    
    # Use pd.concat with proper handling
    tdoc_df = pd.DataFrame([tdoc_row], columns=enhanced_df.columns)
    enhanced_df = pd.concat([enhanced_df, tdoc_df], ignore_index=True)
    
    return enhanced_df

def save_with_formatting(final_df, output_filename, group_info=None):
    """Save DataFrame with Excel formatting - thick borders around each group perimeter"""
    try:
        # Save as Excel file first
        excel_filename = output_filename.replace('.csv', '.xlsx')
        final_df.to_excel(excel_filename, index=False)
        
        # Open workbook for formatting
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active
        
        # Define fills
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        orange_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")  # Orange Accent 2
        
        # Define thick border styles for different positions
        thick_top = Border(top=Side(style='thick'))
        thick_bottom = Border(bottom=Side(style='thick'))
        thick_left = Border(left=Side(style='thick'))
        thick_right = Border(right=Side(style='thick'))
        thick_top_left = Border(top=Side(style='thick'), left=Side(style='thick'))
        thick_top_right = Border(top=Side(style='thick'), right=Side(style='thick'))
        thick_bottom_left = Border(bottom=Side(style='thick'), left=Side(style='thick'))
        thick_bottom_right = Border(bottom=Side(style='thick'), right=Side(style='thick'))
        
        # Find Step column index
        step_col_index = None
        for col_idx, cell in enumerate(ws[1]):
            if cell.value == 'Step':
                step_col_index = col_idx + 1  # openpyxl uses 1-based indexing
                break
        
        # Apply formatting per group
        if step_col_index:
            current_row = 2  # Start after header
            
            while current_row <= ws.max_row:
                # Look for start of group ("1 Beginning On Hand")
                step_value = ws.cell(row=current_row, column=step_col_index).value
                
                if step_value == "1 Beginning On Hand":
                    group_start = current_row
                    group_end = current_row
                    
                    # Find end of group ("9 TDOC")
                    temp_row = current_row + 1
                    while temp_row <= ws.max_row:
                        temp_step_value = ws.cell(row=temp_row, column=step_col_index).value
                        if temp_step_value == "9 TDOC":
                            group_end = temp_row
                            break
                        temp_row += 1
                    
                    # Apply thick border around the perimeter of the group
                    for row in range(group_start, group_end + 1):
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            
                            # Determine which borders to apply based on position
                            is_top = (row == group_start)
                            is_bottom = (row == group_end)
                            is_left = (col == 1)
                            is_right = (col == ws.max_column)
                            
                            # Apply appropriate border combination
                            if is_top and is_left:
                                cell.border = thick_top_left
                            elif is_top and is_right:
                                cell.border = thick_top_right
                            elif is_bottom and is_left:
                                cell.border = thick_bottom_left
                            elif is_bottom and is_right:
                                cell.border = thick_bottom_right
                            elif is_top:
                                cell.border = thick_top
                            elif is_bottom:
                                cell.border = thick_bottom
                            elif is_left:
                                cell.border = thick_left
                            elif is_right:
                                cell.border = thick_right
                            # Interior cells get no border
                    
                    # Apply colors to specific rows within the group (after borders)
                    for row in range(group_start, group_end + 1):
                        step_value = ws.cell(row=row, column=step_col_index).value
                        
                        if step_value in ['Simulated demand', 'Simulated Order']:
                            # Yellow highlighting
                            for col in range(1, ws.max_column + 1):
                                cell = ws.cell(row=row, column=col)
                                cell.fill = yellow_fill
                                
                        elif step_value in ['7 Ending Inventory', '8 DOC - End of Month']:
                            # Orange highlighting
                            for col in range(1, ws.max_column + 1):
                                cell = ws.cell(row=row, column=col)
                                cell.fill = orange_fill
                    
                    # Move to next group
                    current_row = group_end + 1
                else:
                    current_row += 1
        
        # Save formatted workbook
        wb.save(excel_filename)
        print(f"✅ Formatted Excel file saved as: {excel_filename}")
        
        # Also save CSV version without formatting
        final_df.to_csv(output_filename, index=False)
        print(f"✅ CSV file saved as: {output_filename}")
        
    except Exception as e:
        print(f"⚠️ Error applying Excel formatting: {e}")
        # Fallback to CSV only
        final_df.to_csv(output_filename, index=False)
        print(f"✅ CSV file saved as: {output_filename}")

def process_with_target_days(mrp_filename):
    """Process MRP file with target days to create final report with enhanced logic"""
    print(f"\n🔄 STEP 2: Processing {mrp_filename} with target days...")
    
    try:
        # Load data
        mrp_df = pd.read_csv(mrp_filename)
        
        # Get Dependent Forecast data
        dependent_forecast_df = get_dependent_forecast_data()
        
        # Check if target_days.xlsx exists
        if not os.path.exists("target_days.xlsx"):
            print("⚠️ target_days.xlsx not found. Please ensure the file is in the current directory.")
            return
        
        td = pd.read_excel("target_days.xlsx")
        
        print(f"📊 Loaded MRP data: {len(mrp_df)} records")
        print(f"📊 Loaded target days: {len(td)} records")
        print(f"📊 Loaded Dependent Forecast: {len(dependent_forecast_df)} records")
        
        # Identify columns in target_days
        item_col = find_col(td.columns, ["ItemCode", "Item Code", "Item"], ["item"])
        site_col = find_col(td.columns, ["SiteCode", "Site Code", "Site"], ["site", "warehouse", "location"])
        days_col = find_col(td.columns, ["# of Days", "Days", "Target Days"], ["days"])
        
        print(f"📋 Target days columns - Item: {item_col}, Site: {site_col}, Days: {days_col}")
        
        # Normalize keys in both files
        mrp_df["_ITEM_KEY"] = mrp_df["ItemCode"].map(normalize_key)
        mrp_df["_SITE_KEY"] = mrp_df["SiteCode"].map(normalize_key)
        td["_ITEM_KEY"] = td[item_col].map(normalize_key)
        td["_SITE_KEY"] = td[site_col].map(normalize_key)
        
        # Reduce target_days to unique keys
        td_reduced = (
            td[["_ITEM_KEY", "_SITE_KEY", days_col]]
            .dropna(subset=[days_col])
            .drop_duplicates(subset=["_ITEM_KEY", "_SITE_KEY"], keep="first")
        )
        
        # Create lookup dict
        target_map = td_reduced.set_index(["_ITEM_KEY", "_SITE_KEY"])[days_col].to_dict()
        
        # Month columns are everything after "Step"
        if "Step" not in mrp_df.columns:
            print("⚠️ 'Step' column not found in MRP data. Cannot identify month columns.")
            return
        
        month_columns = list(mrp_df.columns[mrp_df.columns.get_loc("Step") + 1:])
        print(f"📅 Found {len(month_columns)} month columns: {month_columns[:5]}...")  # Show first 5
        
        # Build final DataFrame with enhanced logic
        final_parts = []
        fallback_used = []  # collect missing matches
        current_row = 2  # Starting row in Excel (after header)
        group_info = []  # Track group boundaries for formatting
        
        print(f"\n🔄 Processing {len(mrp_df.groupby(['_ITEM_KEY', '_SITE_KEY']))} unique item-site combinations...")
        
        for (item_key, site_key), group in mrp_df.groupby(["_ITEM_KEY", "_SITE_KEY"], sort=False):
            if (item_key, site_key) in target_map:
                days_value = target_map[(item_key, site_key)]
            else:
                days_value = 120
                fallback_used.append((group.iloc[0]["ItemCode"], group.iloc[0]["SiteCode"]))
            
            group_start_row = current_row
            
            # Create enhanced rows with correct formulas based on current row position
            enhanced_group = create_enhanced_rows(group, month_columns, days_value, current_row, dependent_forecast_df)
            final_parts.append(enhanced_group)
            
            # Update current row for next group
            current_row += len(enhanced_group)
            group_end_row = current_row - 1
            
            # Track group boundaries for formatting
            group_info.append((group_start_row, group_end_row))
        
        final_df = pd.concat(final_parts, ignore_index=True)
        
        # Remove helper columns
        final_df = final_df.drop(columns=["_ITEM_KEY", "_SITE_KEY"])
        
        # Save file with today's date and formatting
        today = datetime.now().strftime("%Y-%m-%d")
        output_filename = f"MRP with Target Days_{today}.csv"
        save_with_formatting(final_df, output_filename, group_info)
        
        # Reporting
        print(f"\n✅ Enhanced final file saved with formatting!")
        print(f"📊 Final records: {len(final_df)}")
        print(f"📊 Unique item-site combinations: {len(final_parts)}")
        print(f"🆕 Added features:")
        print(f"   ✅ Dependent Forecast rows with improved month matching")
        print(f"   ✅ Simulated demand rows (after 2 Demand)")
        print(f"   ✅ Simulated Order rows (after 6 Suggested PO Receipt)")
        print(f"   ✅ Formula-based Ending Inventory calculations with Dependent Forecast")
        print(f"   ✅ Formula-based DOC - End of Month calculations with Dependent Forecast")
        print(f"   ✅ Beginning On Hand chain formulas (F2=E10, G2=F10, etc.)")
        print(f"   🎨 Yellow highlighting for Simulated rows")
        print(f"   🎨 Orange highlighting for calculation rows")
        print(f"   📦 Thick borders per group (not per row type)")
        
        if fallback_used:
            print(f"⚠️ {len(fallback_used)} combinations used fallback 75 because no match was found in target_days.xlsx:")
            for item, site in fallback_used[:10]:  # Show first 10
                print(f"  - ItemCode: {item}, SiteCode: {site}")
            if len(fallback_used) > 10:
                print(f"  ... and {len(fallback_used) - 10} more")
        else:
            print("🎉 All combinations matched target_days.xlsx without fallback.")
            
    except Exception as e:
        print(f"❌ Error processing target days: {e}")
        import traceback
        traceback.print_exc()

# --------- Main Function ---------

def main():
    print("🚀 ENHANCED StockIQ MRP Report Generator with Target Days")
    print("="*70)
    print("STEP 1: Fetch data from StockIQ API and create MRP CSV")
    print("STEP 2: Process MRP CSV with enhanced logic:")
    print("        ✅ Add Dependent Forecast rows with better month matching")
    print("        ✅ Add Simulated Demand rows")
    print("        ✅ Add Simulated Order rows")
    print("        ✅ Formula-based Ending Inventory")
    print("        ✅ Formula-based DOC calculations")
    print("        ✅ Beginning On Hand chain formulas")
    print("        🎨 Excel formatting with colors and thick borders per group")
    print("="*70)
    
    # STEP 1: Get the report from API
    print("\n" + "🔵" * 20 + " STEP 1 " + "🔵" * 20)
    df = get_stockiq_custom_report()
    
    if df is not None:
        # Apply filters with detailed analysis
        filtered_df = apply_filters(df)
        
        if len(filtered_df) > 0:
            # Save the MRP report
            mrp_filename = save_mrp_csv(filtered_df)
            
            if mrp_filename:
                # STEP 2: Process with target days and enhanced logic
                print("\n" + "🟢" * 20 + " STEP 2 " + "🟢" * 20)
                process_with_target_days(mrp_filename)
                today = datetime.now().strftime("%Y-%m-%d")
                print(f"\n🎉 ENHANCED PROCESS COMPLETED!")
                print(f"📁 Check these files for results:")
                print(f"   📊 'MRP with Target Days_{today}.csv' (data file)")
                print(f"   🎨 'MRP with Target Days_{today}.xlsx' (formatted Excel file)")
            else:
                print("\n❌ Failed to save MRP file, cannot proceed to Step 2")
        else:
            print("\n❌ No data to process after filtering.")
    else:
        print("❌ Failed to retrieve report data from API")

if __name__ == "__main__":
    main()