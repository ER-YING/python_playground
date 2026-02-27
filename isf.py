from openpyxl import load_workbook

def fill_isf_form(filename, data):
    wb = load_workbook(filename)
    ws = wb.active  # Assumes that the data is in the first sheet

    # Purchase Order Numbers
    po_numbers = data['purchase_order_number'].split(';')
    for index, po in enumerate(po_numbers):
        ws[f'{chr(71+index)}5'] = po.strip()  # 'G5' is ASCII 71

    # Sailing Date
    ws['G6'] = data['sailing_date']

    # MBL and Scace Code
    ws['G8'] = data['mbl_scace_code']

    # Optional Container #
    if 'container_number' in data:
        ws['G9'] = data['container_number']

    # Optional Vessel #
    if 'vessel_number' in data:
        ws['G10'] = data['vessel_number']

    # AMS House Bills of Lading #
    ws['G13'] = data['ams_house_bills']

    # HBL SCAC Code
    ws['D14'] = data['hbl_scac_code']

    # Manufacturer details
    manufacturer = data['manufacturer']
    if manufacturer == "AHC":
        for cell in ['H18', 'H30', 'H42', 'H48']:
            ws[cell] = "AHC LTD."
        for cell in ['H19', 'H31', 'H43', 'H49']:
            ws[cell] = "189,HAIZHEN RD,DUSHANGANG TOWN, PINGHU"
        for cell in ['H21', 'H33', 'H45', 'H51']:
            ws[cell] = "JIAXING"
        for cell in ['H22', 'H34', 'H46', 'H52']:
            ws[cell] = "ZHEJIANG"
        for cell in ['H23', 'H35', 'H47', 'H53']:
            ws[cell] = "CHINA"
        ws['G25'] = "3401.19"
    elif manufacturer == "Yancheng":
        for addr in ['H18', 'H42']:
            ws[addr] = "Yancheng Chong Jia Technology Co., Ltd"
        for addr in ['H19', 'H43']:
            ws[addr] = "West side of youdian road, Qinfeng mechanical and electrical industrial park,"
        for city in ['H20', 'H44']:
            ws[city] = "Jianhu town"
        for province in ['H21', 'H45']:
            ws[province] = "Yan Cheng"
        for province in ['H22', 'H46']:
            ws[province] = "Jiang Su"
        for country in ['H23', 'H29', 'H35', 'H47', 'H53']:
            ws[country] = "CHINA"
        ws['G25'] = "3923.21"
        # Additional Manufacturer "Sunner World Trading CO., Ltd."
        for addr in ['H30', 'H48']:
            ws[addr] = "Sunner World Trading CO., Ltd."
        for addr in ['H31', 'H49']:
            ws[addr] = "Suite 706-707 Huitu Mansion 898 Tong Pu Road"
        for city in ['H33', 'H51', 'H34', 'H52']:
            ws[city] = "Shanghai"
    elif manufacturer == "Thai Duong":
        for cell in ['H18', 'H30', 'H42', 'H48']:
            ws[cell] = "THAI DUONG RUBBER JOINT STOCK COMPANY"
        for cell in ['H19', 'H31', 'H43', 'H49']:
            ws[cell] = "LOT 4, ROAD 7, TAN TAO INDUSTRIAL PARK, BINH TAN DISTRICT"
        for cell in ['H21', 'H33', 'H45', 'H51']:
            ws[cell] = "HO CHI MINH CITY"
        for cell in ['H22', 'H34', 'H46', 'H52']:
            ws[cell] = "HO CHI MINH CITY"
        for cell in ['H23', 'H35', 'H47', 'H53']:
            ws[cell] = "VIETNAM"
        ws['G25'] = "4016.99"
    elif manufacturer == "Tianjin":
        for cell in ['H18', 'H30', 'H42', 'H48']:
            ws[cell] = "TIANJIN YIYI HYGIENE PRODUCTS CO., LTD."
        for cell in ['H19', 'H31', 'H43', 'H49']:
            ws[cell] = "ZHANGJIAWO INDUSTRIAL PARK,XIQING ECONOMIC DEVELOPMENT AREA"
        for cell in ['H21', 'H33', 'H45', 'H51']:
            ws[cell] = "TIAN JIN"
        for cell in ['H22', 'H34', 'H46', 'H52']:
            ws[cell] = "TIAN JIN"
        for cell in ['H23', 'H35', 'H47', 'H53']:
            ws[cell] = "CHINA"
        ws['G25'] = "4818.90"
    # Destination Warehouse
    destination_warehouse = data['destination_warehouse']
    if destination_warehouse == "Source NJ":
        ws['H54'] = "SOURCE NEW JERSEY"
        ws['H55'] = "2 MIDDLESEX AVENUE, SUITE A"
        ws['H57'] = "MONROE"
        ws['H58'] = "NJ"
    elif destination_warehouse == "Source MTB":
        ws['H54'] = "SOURCE LOGISTICS LOS ANGELES, LLC"
        ws['H55'] = "812 UNION ST"
        ws['H56'] = "MONTEBELLO"
        ws['H57'] = "CA"

        # Create the new filename based on multiple Purchase Order Numbers
    if len(po_numbers) > 1:
        new_filename = '+'.join(po_numbers) + ' ISF.xlsx'
    else:
        new_filename = po_numbers[0] + ' ISF.xlsx'

    # Save the workbook with the new filename
    wb.save(new_filename)


# Example data dictionary to fill the form
data = {
    'purchase_order_number': 'ER3983; ER3778; ER4316; ER4336; ER4103',
    'sailing_date': '2025-10-02',
    'mbl_scace_code': 'COSU6430988590',
    'container_number': 'CSNU8740805',
    'vessel_number': 'COSCO EUROPE 108E',
    'ams_house_bills': 'SHA251106700',
    'hbl_scac_code': 'STIW',
    'manufacturer': 'Yancheng',
    'destination_warehouse': 'Source MTB'
}

fill_isf_form('../Playwright_expercice/ISF_Python.xlsx', data)
