import gspread
import pandas as pd
import os
import io
import pickle
import base64
from datetime import timedelta
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# --- Third-party libraries ---
from openpyxl import load_workbook
from gspread_dataframe import get_as_dataframe
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage

# ==============================================================================
# SCRIPT CONFIGURATION - - MODIFY THIS SECTION
# ==============================================================================

INTERNAL_IDS_TO_PROCESS = ['2082']
CLIENT_SECRETS_FILE = 'client_secret_445547975943-l4bvv3advto71sovc6rn5ndai1gmhib9.apps.googleusercontent.com.json'
PO_CHECK_SPREADSHEET_ID = '1DDb_Z08ET94wdmD6a93UPVgWQ73y0F2tQIlgHqMiFCc'
SERVICE_ACCOUNT_FILE = 'google_drive_sp_key.json'
INBOUND_SHIPMENT_TRACKER_TAB = 'Inbound Shipment Tracker'
SENDER_EMAIL = "ying@earthrated.com"
TO_RECIPIENTS = ["ying@earthrated.com"]
CC_RECIPIENTS = ["ying@earthrated.com"]
ISF_TEMPLATE_FILENAME = 'ISF_Python.xlsx'
FC_ADDRESS_LOCAL_FILE = 'FC ADDRESS.xlsx'

# --- Signature Configuration ---
SIGNATURE_IMAGE_FILENAME = 'signature_logo.png'


# --- Automated Folder Search Configuration ---
OPERATIONS_SHARED_DRIVE_ID = "0AB9VXgYGzJL6Uk9PVA" # <-- PASTE ID of "Operations" Shared Drive

# --- Supplier Name for Special Case ---
TIANJIN_YIYI_SUPPLIER_NAME = "Tianjin Yiyi Hygiene Products Co., Ltd."

# ==============================================================================
# AUTHENTICATION & CORE FUNCTIONS
# ==============================================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SHEETS_SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']
DRIVE_SCOPES = ['https://www.googleapis.com/auth/drive'] # Changed to read/write

def get_google_creds(token_name, scopes):
    token_path = os.path.join(SCRIPT_DIR, token_name)
    creds = None
    if os.path.exists(token_path):
        with open(token_path, 'rb') as token: creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, scopes)
            creds = flow.run_local_server(port=0)
        with open(token_path, 'wb') as token: pickle.dump(creds, token)
    return creds

def get_google_service_account_creds(credentials_file, scopes):
    from google.oauth2 import service_account
    return service_account.Credentials.from_service_account_file(credentials_file, scopes=scopes)

def fill_isf_form(filename, data):
    # This function is correct and unchanged
    wb = load_workbook(filename)
    ws = wb.active
    po_numbers = data.get('purchase_order_number', '').split(';')
    for index, po in enumerate(po_numbers):
        if po: ws[f'{chr(71+index)}5'] = po.strip()
    ws['G6'] = data.get('sailing_date', '')
    ws['G8'] = data.get('mbl_scace_code', '')
    ws['G9'] = data.get('container_number', '')
    ws['G10'] = data.get('vessel_number', '')
    ws['G13'] = data.get('ams_house_bills', '')
    ws['D14'] = data.get('hbl_scac_code', '')
    manufacturer = data.get('manufacturer', '')
    if manufacturer == "American Hygienics Corporation":
        for cell in ['H18', 'H30', 'H42', 'H48']: ws[cell] = "AHC LTD."
        for cell in ['H19', 'H31', 'H43', 'H49']: ws[cell] = "189,HAIZHEN RD,DUSHANGANG TOWN, PINGHU"
        for cell in ['H21', 'H33', 'H45', 'H51']: ws[cell] = "JIAXING"
        for cell in ['H22', 'H34', 'H46', 'H52']: ws[cell] = "ZHEJIANG"
        for cell in ['H23', 'H35', 'H47', 'H53']: ws[cell] = "CHINA"
        ws['G25'] = "3401.19"
    elif manufacturer == "SUNNER GROUP CO., LTD.":
        for addr in ['H18', 'H42']: ws[addr] = "Yancheng Chong Jia Technology Co., Ltd"
        for addr in ['H19', 'H43']: ws[addr] = "West side of youdian road, Qinfeng mechanical and electrical industrial park,"
        for city in ['H20', 'H44']: ws[city] = "Jianhu town"
        for province in ['H21', 'H45']: ws[province] = "Yan Cheng"
        for province in ['H22', 'H46']: ws[province] = "Jiang Su"
        for country in ['H23', 'H29', 'H35', 'H47', 'H53']: ws[country] = "CHINA"
        ws['G25'] = "3923.21"
        for addr in ['H30', 'H48']: ws[addr] = "Sunner World Trading CO., Ltd."
        for addr in ['H31', 'H49']: ws[addr] = "Suite 706-707 Huitu Mansion 898 Tong Pu Road"
        for city in ['H33', 'H51', 'H34', 'H52']: ws[city] = "Shanghai"
    elif manufacturer == "Thai Duong Rubber Joint Stock Company":
        for cell in ['H18', 'H30', 'H42', 'H48']: ws[cell] = "THAI DUONG RUBBER JOINT STOCK COMPANY"
        for cell in ['H19', 'H31', 'H43', 'H49']: ws[cell] = "LOT 4, ROAD 7, TAN TAO INDUSTRIAL PARK, BINH TAN DISTRICT"
        for cell in ['H21', 'H33', 'H45', 'H51']: ws[cell] = "HO CHI MINH CITY"
        for cell in ['H22', 'H34', 'H46', 'H52']: ws[cell] = "HO CHI MINH CITY"
        for cell in ['H23', 'H35', 'H47', 'H53']: ws[cell] = "VIETNAM"
        ws['G25'] = "4016.99"
    elif manufacturer == "Tianjin Yiyi Hygiene Products Co., Ltd.":
        for cell in ['H18', 'H30', 'H42', 'H48']: ws[cell] = "TIANJIN YIYI HYGIENE PRODUCTS CO., LTD."
        for cell in ['H19', 'H31', 'H43', 'H49']: ws[cell] = "ZHANGJIAWO INDUSTRIAL PARK,XIQING ECONOMIC DEVELOPMENT AREA"
        for cell in ['H21', 'H33', 'H45', 'H51']: ws[cell] = "TIAN JIN"
        for cell in ['H22', 'H34', 'H46', 'H52']: ws[cell] = "TIAN JIN"
        for cell in ['H23', 'H35', 'H47', 'H53']: ws[cell] = "CHINA"
        ws['G25'] = "4818.90"
    if 'amazon_fc_address' in data and data['amazon_fc_address']:
        address_info = data['amazon_fc_address']
        ws['H54'], ws['H55'], ws['H57'], ws['H58'] = address_info.get('company_name', ''), address_info.get('address', ''), address_info.get('city', ''), address_info.get('state', '')
    else:
        destination_warehouse = data.get('destination_warehouse', '')
        if destination_warehouse == "Source NJ":
            ws['H54'], ws['H55'], ws['H57'], ws['H58'] = "SOURCE NEW JERSEY", "2 MIDDLESEX AVENUE, SUITE A", "MONROE", "NJ"
        elif destination_warehouse == "Source Montebello":
            ws['H54'], ws['H55'], ws['H57'], ws['H58'] = "SOURCE LOGISTICS LOS ANGELES, LLC", "812 UNION ST", "MONTEBELLO", "CA"
    # The filename for email attachment is the user-friendly PO# version
    po_string = '+'.join([po.strip() for po in data.get('purchase_order_number', '').split(';') if po])
    email_filename = f"{po_string} ISF.xlsx"
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return email_filename, output_stream.getvalue()

def send_email_with_attachments(to, cc, subject, html_body, sender, attachments_list, images_to_embed):
    # This function is correct and unchanged
    creds = get_google_creds('token_gmail.pickle', GMAIL_SCOPES)
    service = build('gmail', 'v1', credentials=creds)
    message = MIMEMultipart('related')
    message['to'], message['from'], message['subject'] = ", ".join(to), sender, subject
    if cc: message['cc'] = ", ".join(cc)
    msg_alternative = MIMEMultipart('alternative')
    message.attach(msg_alternative)
    msg_text = MIMEText(html_body, 'html')
    msg_alternative.attach(msg_text)
    for image_path in images_to_embed:
        image_cid = os.path.basename(image_path)
        try:
            with open(image_path, 'rb') as f:
                msg_image = MIMEImage(f.read())
                msg_image.add_header('Content-ID', f'<{image_cid}>')
                message.attach(msg_image)
        except FileNotFoundError: print(f"Warning: Signature image '{image_path}' not found.")
    for name, data in attachments_list:
        part = MIMEApplication(data, Name=name)
        part['Content-Disposition'] = f'attachment; filename="{name}"'
        message.attach(part)
    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
    try:
        sent_message = service.users().messages().send(userId='me', body={'raw': raw_message}).execute()
        print(f"✅ Email '{subject}' with {len(attachments_list)} attachments sent successfully.")
    except Exception as error: print(f"❌ An error occurred while sending email: {error}")

def check_multiple_suppliers(shipment_df):
    """Checks the 'Supplier Name' column for unique, non-empty entries."""
    unique_suppliers = shipment_df['Supplier Name'].dropna().unique()
    return [str(s).strip() for s in unique_suppliers if str(s).strip()]

def get_manufacturer_input(unique_suppliers):
    """Handles the logic for selecting the correct manufacturer, prompting the user if needed."""
    available_manufacturers = [
        "American Hygienics Corporation", "SUNNER GROUP CO., LTD.",
        "Thai Duong Rubber Joint Stock Company", "Tianjin Yiyi Hygiene Products Co., Ltd."
    ]
    
    # Case 1: Exactly one supplier is found in the sheet
    if len(unique_suppliers) == 1:
        print(f"\n✅ Single supplier found: {unique_suppliers[0]}")
        # Check if this supplier is one of our known manufacturers
        for m in available_manufacturers:
            if unique_suppliers[0].lower() == m.lower():
                # If it's a known one, ask the user to confirm
                if input(f"Use this supplier as manufacturer? (Y/n): ").strip().lower() in ['', 'y', 'yes']:
                    return m # <-- If user says yes, we are DONE and exit the function.
                break # Stop checking other manufacturers
    
    # Case 2: No suppliers were found, or multiple suppliers were found,
    # or the user said 'n' to using the single found supplier.
    elif not unique_suppliers:
        print("\n⚠️ WARNING: No supplier name found in the sheet.")
    else: # len > 1
        print(f"\n⚠️ WARNING: Multiple suppliers found for this Internal ID:")
        for i, s in enumerate(unique_suppliers, 1):
            print(f"   {i}. {s}")
    
    # This block will now correctly run for all cases that need manual selection
    print("\nPlease choose the correct manufacturer for the ISF form:")
    for i, m in enumerate(available_manufacturers, 1):
        print(f"  {i}. {m}")
    
    while True:
        choice = input(f"Enter option number (1-{len(available_manufacturers)}) or full name: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(available_manufacturers):
            selected = available_manufacturers[int(choice) - 1]
            print(f"✅ Selected: {selected}")
            return selected
        for m in available_manufacturers:
            if choice.lower() == m.lower():
                print(f"✅ Selected: {m}")
                return m
        print(f"❌ Invalid input. Please enter a valid number or full name.")

    print("\nPlease choose the correct manufacturer:")
    for i, m in enumerate(available_manufacturers, 1): print(f"  {i}. {m}")
    while True:
        choice = input(f"Enter option number (1-{len(available_manufacturers)}): ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(available_manufacturers):
            selected = available_manufacturers[int(choice) - 1]
            print(f"✅ Selected: {selected}"); return selected
        for m in available_manufacturers:
            if choice.lower() == m.lower(): print(f"✅ Selected: {m}"); return m
        print(f"❌ Invalid input.")

def find_shipment_subfolders(main_folder_id, service):
    """Finds the 'Freight' and 'Paperwork' subfolders inside a main shipment folder."""
    folder_ids = {'freight': None, 'paperwork': None}
    for subfolder_name in ['Freight', 'Paperwork']:
        query = f"'{main_folder_id}' in parents and name='{subfolder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        try:
            results = service.files().list(q=query, supportsAllDrives=True, includeItemsFromAllDrives=True, fields="files(id, name)").execute()
            found = results.get('files', [])
            if found: folder_ids[subfolder_name.lower()] = found[0]['id']; print(f"   -> Found '{subfolder_name}' sub-folder.")
            else: print(f"   -> ❌ Could not find a '{subfolder_name}' sub-folder.")
        except Exception as e: print(f"Error finding subfolder '{subfolder_name}': {e}")
    return folder_ids

def download_drive_files(parent_folder_id, file_names_to_find, service):
    """Downloads specified files from a given parent folder ID on Google Drive."""
    downloaded_files = {}
    if not parent_folder_id: return {}
    try:
        query = f"'{parent_folder_id}' in parents and trashed=false"
        results = service.files().list(q=query, fields="files(id, name)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        files_in_folder = results.get('files', [])
        for file_to_find in file_names_to_find:
            found_match = False
            for file in files_in_folder:
                if file['name'].lower() == file_to_find.lower():
                    print(f"   -> Downloading '{file['name']}'...")
                    request = service.files().get_media(fileId=file['id'], supportsAllDrives=True)
                    file_io = io.BytesIO()
                    downloader = MediaIoBaseDownload(file_io, request)
                    done = False
                    while not done: status, done = downloader.next_chunk()
                    file_io.seek(0)
                    downloaded_files[file['name']] = file_io.getvalue()
                    found_match = True
                    break
            if not found_match: print(f"   -> ⚠️  Warning: File '{file_to_find}' not found.")
    except Exception as e: print(f"   -> ❌ Error downloading files: {e}")
    return downloaded_files

# --- NEW: Function to upload a file to a specific Drive folder ---
def upload_to_drive(parent_folder_id, filename, file_data, service):
    """Uploads file data to a specified Google Drive folder."""
    print(f"Uploading '{filename}' to Google Drive...")
    file_metadata = {'name': filename, 'parents': [parent_folder_id]}
    media = MediaIoBaseUpload(io.BytesIO(file_data), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    try:
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id',
            supportsAllDrives=True
        ).execute()
        print(f"   -> ✅ Successfully uploaded file. File ID: {file.get('id')}")
        return True
    except Exception as e:
        print(f"   -> ❌ An error occurred during file upload: {e}")
        return False

# --- NEW: Function to automatically find the main shipment folder ---
def find_main_shipment_folder(internal_id, drive_id, service):
    """Searches the entire Shared Drive for the main shipment folder."""
    if not drive_id or drive_id == "YOUR_OPERATIONS_SHARED_DRIVE_ID_HERE":
        print("❌ FATAL ERROR: OPERATIONS_SHARED_DRIVE_ID is not set.")
        return None
    try:
        print(f"Searching for main shipment folder containing 'INBSHIP{internal_id}'...")
        query = f"name contains 'INBSHIP{internal_id}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        results = service.files().list(q=query, corpora="drive", driveId=drive_id,
                                       includeItemsFromAllDrives=True, supportsAllDrives=True,
                                       fields="files(id, name)").execute()
        folders = results.get('files', [])
        if not folders: print(f"   -> ❌ No main shipment folder found for Internal ID '{internal_id}'."); return None
        folder = folders[0]
        print(f"   -> Found main shipment folder: '{folder['name']}'")
        return folder['id']
    except Exception as e:
        print(f"   -> ❌ An error occurred while searching for Drive folder: {e}")
        return None


# ==============================================================================
# MAIN EXECUTION SCRIPT
# ==============================================================================
if __name__ == "__main__":
    print("Starting ISF Automation Script...")

    # Initialize services once
    drive_creds = get_google_service_account_creds(SERVICE_ACCOUNT_FILE, DRIVE_SCOPES)
    drive_service = build('drive', 'v3', credentials=drive_creds)

    try:
        # Load all data from Google Sheets and local files first
        sheets_creds = get_google_creds('token_sheets.pickle', SHEETS_SCOPES)
        client = gspread.authorize(sheets_creds)
        sheet = client.open_by_key(PO_CHECK_SPREADSHEET_ID)
        worksheet = sheet.worksheet(INBOUND_SHIPMENT_TRACKER_TAB)
        df = get_as_dataframe(worksheet, evaluate_formulas=True)
        df['Internal Id'] = df['Internal Id'].astype(str).str.strip().replace(r'\.0$', '', regex=True)
        print(f"Successfully loaded data from '{INBOUND_SHIPMENT_TRACKER_TAB}'.")
        
        expanded_fc_path = os.path.expanduser(FC_ADDRESS_LOCAL_FILE)
        fc_address_df = pd.read_excel(expanded_fc_path)
        fc_address_df['CITY_NORMALIZED'] = fc_address_df['CITY'].str.strip().str.upper()
        
    except Exception as e:
        print(f"❌ FATAL ERROR during data loading: {e}"); exit()

    # Prepare signature HTML once, before the loop starts
    signature_html = f"""
    <br><br>
    <div><table cellpadding="0" cellspacing="0" border="0"><tr>
    <td valign="top" style="padding-right: 15px;"><img src="cid:{SIGNATURE_IMAGE_FILENAME}" alt="Earth Rated Logo" width="120"></td>
    <td valign="top" style="color: #333333; border-left: 1px solid #dcdcdc; padding-left: 15px; font-size: 10pt;">
    <b style="color: #00573c;">Ying Ji</b><br>Import & Export Specialist<br>Spécialiste, import et export<br><br>1.888.354.2818<br>earthrated.com<br><br>
    </td></tr></table>
    <p style="font-size: 10pt; color: #00573c; font-weight: bold; margin-top: 15px;">Proudly B Corp Certified<br>Entreprise certifiée B Corp</p></div>
    """

    for internal_id in INTERNAL_IDS_TO_PROCESS:
        print(f"\n{'='*25}\n--- Processing Internal ID: {internal_id} ---\n{'='*25}")

        shipment_df = df[df['Internal Id'] == internal_id].copy()
        if shipment_df.empty:
            print(f"Warning: Internal ID {internal_id} not found. Skipping."); continue

        first_row = shipment_df.iloc[0]
        destination_warehouse = str(first_row.get('Location', ''))
        supplier_name = str(first_row.get('Supplier Name', ''))

        # Step 1: Find the main shipment folder in Google Drive
        main_shipment_folder_id = find_main_shipment_folder(internal_id, OPERATIONS_SHARED_DRIVE_ID, drive_service)
        if not main_shipment_folder_id:
            print(f"❌ Critical folder missing in Drive for Internal ID {internal_id}. Cannot proceed. Skipping.")
            continue

        # Step 2: Generate ISF file in memory
        selected_manufacturer = get_manufacturer_input(check_multiple_suppliers(shipment_df))
        po_numbers = ";".join(shipment_df['PO#'].astype(str).unique())
        mbl_number = str(first_row.get('MBL#', ''))
        ams_house_bills, hbl_scac_code = "", ""
        if not mbl_number.upper().startswith('COSU'):
            ams_house_bills = str(first_row.get('Booking#', ''))
            if ams_house_bills and len(ams_house_bills) >= 4: hbl_scac_code = ams_house_bills[:4]
        
        isf_data = {'purchase_order_number': po_numbers, 'sailing_date': pd.to_datetime(first_row['ETD']).strftime('%Y-%m-%d') if pd.notna(first_row['ETD']) else '', 'mbl_scace_code': mbl_number, 'container_number': str(first_row.get('Container#', '')), 'vessel_number': '', 'ams_house_bills': ams_house_bills, 'hbl_scac_code': hbl_scac_code, 'manufacturer': selected_manufacturer, 'destination_warehouse': destination_warehouse}

        if destination_warehouse == "Golden State FC LLC (AGL)":
            fc_city_input = input("Location is AGL. Please input Amazon FC city name: ")
            match = fc_address_df[fc_address_df['CITY_NORMALIZED'] == fc_city_input.strip().upper()]
            if not match.empty:
                address_row = match.iloc[0]
                isf_data['amazon_fc_address'] = {'company_name': address_row.get('COMPANY_NAME'), 'address': address_row.get('ADDRESS'), 'city': address_row.get('CITY'), 'state': address_row.get('STATES')}
            else:
                print(f"❌ ERROR: Could not find FC address for city '{fc_city_input}'. Skipping."); continue
        
        try:
            email_filename, attachment_data = fill_isf_form(ISF_TEMPLATE_FILENAME, isf_data)
        except Exception as e:
            print(f"❌ Error generating ISF form: {e}"); continue

        # --- NEW: Find sub-folders BEFORE uploading ---
        subfolder_ids = find_shipment_subfolders(main_shipment_folder_id, drive_service)
        if not subfolder_ids or not subfolder_ids['freight']:
             print(f"❌ Critical 'Freight' sub-folder not found. Cannot upload ISF or find attachments. Skipping.")
             continue
        
        # Step 3: Upload the generated ISF file to Google Drive
        drive_isf_filename = f"INBSHIP{internal_id} ISF.xlsx"
        upload_to_drive(subfolder_ids['freight'], drive_isf_filename, attachment_data, drive_service)

        # Step 4: Prepare attachments for the email based on new rules
        attachments_list = [(email_filename, attachment_data)]
        subfolder_ids = find_shipment_subfolders(main_shipment_folder_id, drive_service)

        docs_to_download = [(subfolder_ids['freight'], [f"INBSHIP{internal_id}BOL.pdf"])]
        
        if destination_warehouse == "Golden State FC LLC (AGL)":
            if supplier_name == TIANJIN_YIYI_SUPPLIER_NAME:
                print("   -> Special Case: 'Tianjin Yiyi' AGL shipment. Requiring IV and PL paperwork.")
                docs_to_download.append((subfolder_ids['paperwork'], [f"INBSHIP{internal_id}PaperworkIV.pdf", f"INBSHIP{internal_id}PaperworkPL.pdf"]))
            else:
                print("   -> Standard AGL shipment. Requiring standard Paperwork.")
                docs_to_download.append((subfolder_ids['paperwork'], [f"INBSHIP{internal_id}Paperwork.pdf"]))

        for folder_id, file_list in docs_to_download:
            downloaded = download_drive_files(folder_id, file_list, drive_service)
            attachments_list.extend(list(downloaded.items()))

        # Step 5: Construct and send the email
        shipping_port = {"American Hygienics Corporation": "SHA", "SUNNER GROUP CO., LTD.": "SHA", "Tianjin Yiyi Hygiene Products Co., Ltd.": "XNG", "Thai Duong Rubber Joint Stock Company": "VNM"}.get(selected_manufacturer, "N/A")
        etd_date_str = pd.to_datetime(first_row['ETD']).strftime('%-m/%-d') if pd.notna(first_row['ETD']) else 'N/A'
        eta_date_str = 'N/A'
        if pd.notna(first_row.get('ETA')):
            try: eta_date_str = pd.to_datetime(first_row['ETA']).strftime('%-m/%-d')
            except: pass
        location = str(first_row.get('Location', 'N/A'))
        
        email_subject = f"EARTH RATED ISF filing - INBSHIP{internal_id} - ETD {shipping_port} {etd_date_str} ETA {location} {eta_date_str}"
        
        # Generate HTML with conditional message
        additional_message = ""
        if destination_warehouse == "Golden State FC LLC (AGL)":
            additional_message = "<p><b>This is Amazon FBA container, carrier will send you AN directly, I attached other shipping docs for the entry</b></p>"
        html_body = f"""
        <html><body>
            <p>Hi Team,</p><p>Good Day!</p>
            <p>Pls help to file ISF with for subject PO.</p>
            {additional_message}
            <p>Let me know if there is any issues.</p>
            <p>Thank you!</p>
            {signature_html} 
        </body></html>"""
        
        images_to_embed = [os.path.join(SCRIPT_DIR, SIGNATURE_IMAGE_FILENAME)]

        send_email_with_attachments(to=TO_RECIPIENTS, cc=CC_RECIPIENTS, subject=email_subject, html_body=html_body, sender=SENDER_EMAIL, attachments_list=attachments_list, images_to_embed=images_to_embed)

    print("\n--- Script finished. ---")