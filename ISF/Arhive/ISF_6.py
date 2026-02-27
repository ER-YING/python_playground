import gspread
import pandas as pd
import os
import io
import pickle
import base64
from datetime import timedelta
from googleapiclient.http import MediaIoBaseDownload

# --- Third-party libraries ---
from openpyxl import load_workbook
from gspread_dataframe import get_as_dataframe
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage

# ==============================================================================
# SCRIPT CONFIGURATION - - MODIFY THIS SECTION
# ==============================================================================

INTERNAL_IDS_TO_PROCESS = ['2050']
CLIENT_SECRETS_FILE = 'client_secret_445547975943-l4bvv3advto71sovc6rn5ndai1gmhib9.apps.googleusercontent.com.json'
PO_CHECK_SPREADSHEET_ID = '1DDb_Z08ET94wdmD6a93UPVgWQ73y0F2tQIlgHqMiFCc'
SERVICE_ACCOUNT_FILE = 'google_drive_sp_key.json' 
INBOUND_SHIPMENT_TRACKER_TAB = 'Inbound Shipment Tracker'
SENDER_EMAIL = "michelle@earthrated.com"
TO_RECIPIENTS = ["customs.phl@savinodelbene.com, kadyn.pease@savinodelbene.com"]
CC_RECIPIENTS = ["michelle@earthrated.com, mike.rateike@savinodelbene.com"]
ISF_TEMPLATE_FILENAME = 'ISF_Python.xlsx'
ISF_OUTPUT_FOLDER = '~/Desktop/coding folder/ISF/'
FC_ADDRESS_LOCAL_FILE = '~/Desktop/coding folder/ISF/FC ADDRESS.xlsx'

# --- Signature Configuration ---
SIGNATURE_IMAGE_FILENAME = 'signature_logo.png'


# ==============================================================================
# AUTHENTICATION & CORE FUNCTIONS (Do not modify)
# ==============================================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SHEETS_SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']
DRIVE_SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

def get_google_creds(token_name, scopes):
    token_path = os.path.join(SCRIPT_DIR, token_name)
    creds = None
    if os.path.exists(token_path):
        with open(token_path, 'rb') as token: creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, scopes)
            creds = flow.run_local_server(port=0)
        with open(token_path, 'wb') as token: pickle.dump(creds, token)
    return creds

def get_google_service_account_creds(credentials_file, scopes):
    """Get credentials using service account JSON file"""
    from google.oauth2 import service_account
    
    creds = service_account.Credentials.from_service_account_file(
        credentials_file, scopes=scopes)
    return creds

def fill_isf_form(filename, data):
    wb = load_workbook(filename)
    ws = wb.active
    po_numbers = data.get('purchase_order_number', '').split(';')
    for index, po in enumerate(po_numbers):
        if po: ws[f'{chr(71+index)}5'] = po.strip()
    ws['G6'] = data.get('sailing_date', '')
    ws['G8'] = data.get('mbl_scace_code', '')
    ws['G9'] = data.get('container_number', '')
    ws['G10'] = data.get('vessel_number', '')
    ws['G13'] = data.get('ams_house_bills', '')
    ws['D14'] = data.get('hbl_scac_code', '')
    manufacturer = data.get('manufacturer', '')
    if manufacturer == "American Hygienics Corporation":
        for cell in ['H18', 'H30', 'H42', 'H48']: ws[cell] = "AHC LTD."
        for cell in ['H19', 'H31', 'H43', 'H49']: ws[cell] = "189,HAIZHEN RD,DUSHANGANG TOWN, PINGHU"
        for cell in ['H21', 'H33', 'H45', 'H51']: ws[cell] = "JIAXING"
        for cell in ['H22', 'H34', 'H46', 'H52']: ws[cell] = "ZHEJIANG"
        for cell in ['H23', 'H35', 'H47', 'H53']: ws[cell] = "CHINA"
        ws['G25'] = "3401.19"
    elif manufacturer == "SUNNER GROUP CO., LTD.":
        for addr in ['H18', 'H42']: ws[addr] = "Yancheng Chong Jia Technology Co., Ltd"
        for addr in ['H19', 'H43']: ws[addr] = "West side of youdian road, Qinfeng mechanical and electrical industrial park,"
        for city in ['H20', 'H44']: ws[city] = "Jianhu town"
        for province in ['H21', 'H45']: ws[province] = "Yan Cheng"
        for province in ['H22', 'H46']: ws[province] = "Jiang Su"
        for country in ['H23', 'H29', 'H35', 'H47', 'H53']: ws[country] = "CHINA"
        ws['G25'] = "3923.21"
        for addr in ['H30', 'H48']: ws[addr] = "Sunner World Trading CO., Ltd."
        for addr in ['H31', 'H49']: ws[addr] = "Suite 706-707 Huitu Mansion 898 Tong Pu Road"
        for city in ['H33', 'H51', 'H34', 'H52']: ws[city] = "Shanghai"
    elif manufacturer == "Thai Duong Rubber Joint Stock Company":
        for cell in ['H18', 'H30', 'H42', 'H48']: ws[cell] = "THAI DUONG RUBBER JOINT STOCK COMPANY"
        for cell in ['H19', 'H31', 'H43', 'H49']: ws[cell] = "LOT 4, ROAD 7, TAN TAO INDUSTRIAL PARK, BINH TAN DISTRICT"
        for cell in ['H21', 'H33', 'H45', 'H51']: ws[cell] = "HO CHI MINH CITY"
        for cell in ['H22', 'H34', 'H46', 'H52']: ws[cell] = "HO CHI MINH CITY"
        for cell in ['H23', 'H35', 'H47', 'H53']: ws[cell] = "VIETNAM"
        ws['G25'] = "4016.99"
    elif manufacturer == "Tianjin Yiyi Hygiene Products Co., Ltd.":
        for cell in ['H18', 'H30', 'H42', 'H48']: ws[cell] = "TIANJIN YIYI HYGIENE PRODUCTS CO., LTD."
        for cell in ['H19', 'H31', 'H43', 'H49']: ws[cell] = "ZHANGJIAWO INDUSTRIAL PARK,XIQING ECONOMIC DEVELOPMENT AREA"
        for cell in ['H21', 'H33', 'H45', 'H51']: ws[cell] = "TIAN JIN"
        for cell in ['H22', 'H34', 'H46', 'H52']: ws[cell] = "TIAN JIN"
        for cell in ['H23', 'H35', 'H47', 'H53']: ws[cell] = "CHINA"
        ws['G25'] = "4818.90"
    if 'amazon_fc_address' in data and data['amazon_fc_address']:
        address_info = data['amazon_fc_address']
        ws['H54'] = address_info.get('company_name', '')
        ws['H55'] = address_info.get('address', '')
        ws['H57'] = address_info.get('city', '')
        ws['H58'] = address_info.get('state', '')
    else:
        destination_warehouse = data.get('destination_warehouse', '')
        if destination_warehouse == "Source NJ":
            ws['H54'], ws['H55'], ws['H57'], ws['H58'] = "SOURCE NEW JERSEY", "2 MIDDLESEX AVENUE, SUITE A", "MONROE", "NJ"
        elif destination_warehouse == "Source Montebello":
            ws['H54'], ws['H55'], ws['H57'], ws['H58'] = "SOURCE LOGISTICS LOS ANGELES, LLC", "812 UNION ST", "MONTEBELLO", "CA"
    po_string = '+'.join([po.strip() for po in data.get('purchase_order_number', '').split(';') if po])
    new_filename = f"{po_string} ISF.xlsx"
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return new_filename, output_stream.getvalue()

def send_email_with_attachments(to, cc, subject, html_body, sender, attachments_list, images_to_embed):
    """Send email with multiple attachments
    attachments_list: list of tuples [(filename, data), (filename, data), ...]
    """
    creds = get_google_creds('token_gmail.pickle', GMAIL_SCOPES)
    service = build('gmail', 'v1', credentials=creds)
    message = MIMEMultipart('related')
    message['to'] = ", ".join(to)
    if cc: message['cc'] = ", ".join(cc)
    message['from'] = sender
    message['subject'] = subject
    msg_alternative = MIMEMultipart('alternative')
    message.attach(msg_alternative)
    msg_text = MIMEText(html_body, 'html')
    msg_alternative.attach(msg_text)
    
    # Add signature images
    for image_path in images_to_embed:
        image_cid = os.path.basename(image_path)
        try:
            with open(image_path, 'rb') as f:
                msg_image = MIMEImage(f.read())
                msg_image.add_header('Content-ID', f'<{image_cid}>')
                message.attach(msg_image)
        except FileNotFoundError: 
            print(f"Warning: Signature image '{image_path}' not found. Skipping.")
    
    # Add all attachments
    for attachment_name, attachment_data in attachments_list:
        part = MIMEApplication(attachment_data, Name=attachment_name)
        part['Content-Disposition'] = f'attachment; filename="{attachment_name}"'
        message.attach(part)
    
    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
    try:
        sent_message = service.users().messages().send(userId='me', body={'raw': raw_message}).execute()
        all_recipients = to + (cc if cc else [])
        print(f"Email with {len(attachments_list)} attachments sent successfully to {all_recipients}. Message ID: {sent_message['id']}")
    except Exception as error: 
        print(f"An error occurred while sending email: {error}")

# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================
def check_multiple_suppliers(shipment_df):
    unique_suppliers = shipment_df['Supplier Name'].dropna().unique()
    return [str(s).strip() for s in unique_suppliers if str(s).strip()]

def get_manufacturer_input(unique_suppliers):
    available_manufacturers = [
        "American Hygienics Corporation", "SUNNER GROUP CO., LTD.",
        "Thai Duong Rubber Joint Stock Company", "Tianjin Yiyi Hygiene Products Co., Ltd."
    ]
    
    if not unique_suppliers:
        print("\n⚠️  WARNING: No supplier name found in the sheet.")
    elif len(unique_suppliers) > 1:
        print(f"\n⚠️  WARNING: Multiple suppliers found for this Internal ID:")
        for i, supplier in enumerate(unique_suppliers, 1): print(f"   {i}. {supplier}")
    else: # len is 1
        print(f"\n✅ Single supplier found: {unique_suppliers[0]}")
        for manufacturer in available_manufacturers:
            if unique_suppliers[0].lower() == manufacturer.lower():
                use_default = input(f"Use this supplier as manufacturer? (Y/n): ").strip().lower()
                if use_default in ['', 'y', 'yes']: return manufacturer
                break
    
    # This block runs if >1 supplier, no supplier, or user opts not to use the single supplier
    print("\nPlease choose the correct manufacturer for the ISF form:")
    for i, manufacturer in enumerate(available_manufacturers, 1):
        print(f"  {i}. {manufacturer}")
    
    while True:
        choice = input(f"Enter option number (1-{len(available_manufacturers)}) or full name: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(available_manufacturers):
            selected = available_manufacturers[int(choice) - 1]
            print(f"✅ Selected: {selected}")
            return selected
        for manufacturer in available_manufacturers:
            if choice.lower() == manufacturer.lower():
                print(f"✅ Selected: {manufacturer}")
                return manufacturer
        print(f"❌ Invalid input. Please enter a number between 1 and {len(available_manufacturers)} or a full name.")

def download_drive_files(drive_folder_id, file_names_to_find):
    """Final, more reliable function to download specific files from Google Drive."""
    creds = get_google_service_account_creds(SERVICE_ACCOUNT_FILE, DRIVE_SCOPES)
    service = build('drive', 'v3', credentials=creds)
    downloaded_files = {}
    
    try:
        # --- NEW: Get the folder's name with Shared Drive support ---
        folder_metadata = service.files().get(
            fileId=drive_folder_id, 
            fields='name',
            supportsAllDrives=True  # <-- CRITICAL FIX #1
        ).execute()
        print(f"✅ Accessing folder named: '{folder_metadata.get('name')}'")

        # --- NEW: List files with Shared Drive support ---
        query = f"parents in '{drive_folder_id}'"
        results = service.files().list(
            q=query, 
            fields="files(id, name)",
            supportsAllDrives=True,         # <-- CRITICAL FIX #2
            includeItemsFromAllDrives=True  # <-- Also required for listing
        ).execute()
        files_in_folder = results.get('files', [])
        
        if not files_in_folder:
            print("   -> This folder is accessible but no files were found inside.")
            return {}
        
        print(f"   -> Found {len(files_in_folder)} files/subfolders. Searching for required documents...")

        for file_to_find in file_names_to_find:
            found_match = False
            for file in files_in_folder:
                if file_to_find.lower() == file['name'].lower():
                    print(f"   -> Match found: Downloading '{file['name']}'...")
                    request = service.files().get_media(
                        fileId=file['id'],
                        supportsAllDrives=True # <-- CRITICAL FIX #3
                    )
                    file_io = io.BytesIO()
                    downloader = MediaIoBaseDownload(file_io, request)
                    done = False
                    while not done:
                        status, done = downloader.next_chunk()
                    
                    file_io.seek(0)
                    downloaded_files[file['name']] = file_io.getvalue()
                    print(f"      -> Download complete ({len(downloaded_files[file['name']])} bytes).")
                    found_match = True
                    break
            
            if not found_match:
                print(f"   -> ❌ No file found with the exact name '{file_to_find}'")

    except Exception as e:
        print(f"Error listing or downloading files: {e}")
        
    return downloaded_files


# ==============================================================================
# MAIN EXECUTION SCRIPT
# ==============================================================================
if __name__ == "__main__":
    print("Starting ISF automation script...")
    
    try:
        sheets_creds = get_google_creds('token_sheets.pickle', SHEETS_SCOPES)
        client = gspread.authorize(sheets_creds)
        sheet = client.open_by_key(PO_CHECK_SPREADSHEET_ID)
        worksheet = sheet.worksheet(INBOUND_SHIPMENT_TRACKER_TAB)
        df = get_as_dataframe(worksheet, evaluate_formulas=True)
        df['Internal Id'] = df['Internal Id'].astype(str).str.strip().replace(r'\.0$', '', regex=True)
        print(f"Successfully loaded data from '{INBOUND_SHIPMENT_TRACKER_TAB}'.")
        expanded_fc_path = os.path.expanduser(FC_ADDRESS_LOCAL_FILE)
        fc_address_df = pd.read_excel(expanded_fc_path)
        fc_address_df['CITY_NORMALIZED'] = fc_address_df['CITY'].str.strip().str.upper()
        print(f"Successfully loaded data from local FC Address file.")
    except FileNotFoundError:
        print(f"FATAL ERROR: A required file was not found. Please check paths for CLIENT_SECRETS_FILE and FC_ADDRESS_LOCAL_FILE.")
        exit()
    except Exception as e: print(f"Error during data loading: {e}"); exit()

    for internal_id in INTERNAL_IDS_TO_PROCESS:
        print(f"\n--- Processing Internal ID: {internal_id} ---")
        
        shipment_df = df[df['Internal Id'] == internal_id].copy()
        if shipment_df.empty: print(f"Warning: Internal ID {internal_id} not found. Skipping."); continue

        # Interactive Manufacturer Selection
        unique_suppliers = check_multiple_suppliers(shipment_df)
        selected_manufacturer = get_manufacturer_input(unique_suppliers)
        
        first_row = shipment_df.iloc[0]
        isf_data = {}
        destination_warehouse = str(first_row.get('Location', ''))
        isf_data['destination_warehouse'] = destination_warehouse
        
        if destination_warehouse == "Golden State FC LLC (AGL)":
            fc_city_input = input(f"Location is '{destination_warehouse}'. Please input Amazon FC city name: ")
            normalized_city = fc_city_input.strip().upper()
            match = fc_address_df[fc_address_df['CITY_NORMALIZED'] == normalized_city]
            if not match.empty:
                address_row = match.iloc[0]
                isf_data['amazon_fc_address'] = {'company_name': address_row.get('COMPANY_NAME', ''), 'address': address_row.get('ADDRESS', ''), 'city': address_row.get('CITY', ''), 'state': address_row.get('STATES', '')}
                print(f"Found FC Address for '{fc_city_input}': {address_row.get('ADDRESS')}")
            else: print(f"ERROR: Could not find an FC address for city '{fc_city_input}'. Skipping."); continue
        
        po_numbers = ";".join(shipment_df['PO#'].astype(str).unique())
        mbl_number = str(first_row.get('MBL#', ''))
        ams_house_bills, hbl_scac_code = "", ""
        if not mbl_number.upper().startswith('COSU'):
            ams_house_bills = str(first_row.get('Booking#', ''))
            if ams_house_bills and len(ams_house_bills) >= 4: hbl_scac_code = ams_house_bills[:4]
        
        isf_data.update({
            'purchase_order_number': po_numbers,
            'sailing_date': pd.to_datetime(first_row['ETD']).strftime('%Y-%m-%d') if pd.notna(first_row['ETD']) else '',
            'mbl_scace_code': mbl_number, 'container_number': str(first_row.get('Container#', '')),
            'vessel_number': '', 'ams_house_bills': ams_house_bills, 'hbl_scac_code': hbl_scac_code,
            'manufacturer': selected_manufacturer
        })
        print("Successfully extracted data for ISF form.")

        try:
            generated_filename, attachment_data = fill_isf_form(ISF_TEMPLATE_FILENAME, isf_data)
            print(f"Successfully generated ISF file in memory: {generated_filename}")
        except FileNotFoundError: print(f"Error: Template file '{ISF_TEMPLATE_FILENAME}' not found."); continue
        except Exception as e: print(f"Error during ISF form generation: {e}"); continue
            
        if ISF_OUTPUT_FOLDER:
            try:
                expanded_folder_path = os.path.expanduser(ISF_OUTPUT_FOLDER)
                os.makedirs(expanded_folder_path, exist_ok=True)
                output_filepath = os.path.join(expanded_folder_path, generated_filename)
                with open(output_filepath, 'wb') as f: f.write(attachment_data)
                print(f"Successfully saved a copy to: {output_filepath}")
            except Exception as e: print(f"Warning: Could not save local copy. Error: {e}")

        po_subject_part = "+".join(po_numbers.split(';'))
        shipping_port_map = {"American Hygienics Corporation": "SHA", "SUNNER GROUP CO., LTD.": "SHA", "Tianjin Yiyi Hygiene Products Co., Ltd.": "XNG", "Thai Duong Rubber Joint Stock Company": "VNM"}
        shipping_port = shipping_port_map.get(selected_manufacturer, "N/A")
        etd_date_str = pd.to_datetime(first_row['ETD']).strftime('%-m/%-d') if pd.notna(first_row['ETD']) else 'N/A'
        eta_date_str = 'N/A'
        if pd.notna(first_row['Est Delivery Date']):
            eta_date = pd.to_datetime(first_row['Est Delivery Date']) - timedelta(days=5)
            eta_date_str = eta_date.strftime('%-m/%-d')
        location = str(first_row.get('Location', ''))
        email_subject = f"EARTH RATED ISF filing - {internal_id} - {po_subject_part} ETD {shipping_port} {etd_date_str} ETA {location} {eta_date_str}"
        
        # Handle multiple attachments
        attachments_list = [(generated_filename, attachment_data)]  # Start with ISF file

        # Generate HTML with conditional message
        additional_message = ""
        if location == "Golden State FC LLC (AGL)":
            additional_message = "<p style=\"font-family: Arial, sans-serif; font-size: 10pt; color: #333333;\"><b>This is Amazon FBA container, carrier will send you AN directly, I attached other shipping docs for the entry</b></p>"
            
            # Download Drive files
            drive_folder_id = input(f"Please enter Google Drive folder ID for Internal ID {internal_id}: ").strip()
            if drive_folder_id:
                # Search for the full, exact filename instead of a partial name
                files_to_download = ['paperwork.pdf', 'BOL.pdf']
                drive_files = download_drive_files(drive_folder_id, files_to_download)
                for filename, file_data in drive_files.items():
                    attachments_list.append((filename, file_data))

        html_content = f"""
        <html><body>
            <p style="font-family: Arial, sans-serif; font-size: 10pt; color: #333333;">Hi Team,</p>
            <p style="font-family: Arial, sans-serif; font-size: 10pt; color: #333333;">Good Day!</p>
            <p style="font-family: Arial, sans-serif; font-size: 10pt; color: #333333;">Pls help to file ISF for subject PO.</p>
            {additional_message}
            <p style="font-family: Arial, sans-serif; font-size: 10pt; color: #333333;">Let me know if there is any issues.</p>
            <p style="font-family: Arial, sans-serif; font-size: 10pt; color: #333333;">Thanks a lot. Have a great day!</p><br>
            <div style="font-family: Arial, sans-serif; font-size: 10pt;">
            <table cellpadding="0" cellspacing="0" border="0">
            <tr><td valign="top" style="padding-right: 15px;">
            <img src="cid:{SIGNATURE_IMAGE_FILENAME}" alt="Earth Rated Logo" width="120"></td>
            <td valign="top" style="color: #333333; border-left: 1px solid #dcdcdc; padding-left: 15px; font-size: 10pt;">
            <b style="color: #00573c; font-size: 11pt;">Michelle Zhao</b><br>Supply Planner<br>Planificateur de l'approvisionnement<br><br>1.888.354.2818<br>earthrated.com<br><br>
            </td></tr></table></div>
        </body></html>"""

        images_to_embed = [os.path.join(SCRIPT_DIR, SIGNATURE_IMAGE_FILENAME)]
        
        print(f"Preparing to send email with subject: {email_subject}")
        print(f"Total attachments: {len(attachments_list)}")
        send_email_with_attachments(to=TO_RECIPIENTS, cc=CC_RECIPIENTS, subject=email_subject, html_body=html_content, sender=SENDER_EMAIL, attachments_list=attachments_list, images_to_embed=images_to_embed)

    print("\n--- Script finished. ---")